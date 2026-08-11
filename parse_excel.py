# parse_excel.py
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def detect_excel_engine(file_path: str) -> str:
    """Returns 'xlrd' for .xls, 'openpyxl' for .xlsx. Raises ValueError for other extensions."""
    suffix = Path(file_path).suffix.lower()
    if suffix == ".xls":
        return "xlrd"
    if suffix == ".xlsx":
        return "openpyxl"
    raise ValueError(f"Unsupported file extension: {suffix!r}. Expected .xls or .xlsx.")


def load_fc_prefixes(fc_codes_file: str) -> set:
    """Reads an FC codes file, returns set of uppercase prefixes (any length)."""
    prefixes = set()
    try:
        with open(fc_codes_file, "r") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    prefixes.add(line.upper())
    except FileNotFoundError:
        logger.warning(f"FC codes file not found: {fc_codes_file}")
    return prefixes


def load_us_fc_prefixes(us_fc_codes_file: str) -> set:
    """Backward-compatible alias for load_fc_prefixes."""
    return load_fc_prefixes(us_fc_codes_file)


def is_region_fc(fc_code, prefixes: set) -> bool:
    """True if fc_code starts with any known FC prefix from the given set."""
    if not fc_code:
        return False
    fc_str = str(fc_code).strip().upper()
    return any(fc_str.startswith(p) for p in prefixes)


def is_us_fc(fc_code, us_prefixes: set) -> bool:
    """Backward-compatible alias for is_region_fc."""
    return is_region_fc(fc_code, us_prefixes)


def group_by_fba_id(rows: list) -> dict:
    """
    Groups rows by FBA ID. Deduplicates tracking entries.
    Returns: {"FBA123": [{"tracking": "...", "carrier": "..."}, ...]}
    Skips rows with empty/None fba_id.
    If an FBA ID contains "/" with multiple valid IDs (e.g. "STAR-A/STAR-B"),
    each part is treated as a separate FBA ID sharing the same tracking entry.
    """
    result = {}
    for row in rows:
        fba_id_raw = str(row.get("fba_id") or "").strip()
        if not fba_id_raw:
            continue

        # Split combined IDs like "STAR-A/STAR-B" into separate shipments
        parts = [p.strip() for p in fba_id_raw.split("/") if p.strip()]
        if not parts:
            continue  # pure "/" or empty — skip
        # Skip non-Amazon shipment IDs: Walmart (end with "WFA"), TikTok (start with "IBR")
        fba_ids = [
            fba for fba in parts
            if not fba.upper().endswith("WFA") and not fba.upper().startswith("IBR")
        ]
        if not fba_ids:
            continue

        entry = {
            "tracking": str(row.get("tracking_num", "")).strip(),
            "carrier": str(row.get("carrier", "")).strip(),
            "row_number": row.get("row_number"),
        }
        for fba_id in fba_ids:
            if fba_id not in result:
                result[fba_id] = []
            if entry not in result[fba_id]:
                result[fba_id].append(entry)
    return result


def categorize_shipments(grouped: dict) -> tuple:
    """
    Splits grouped FBA shipments into those with usable tracking and those missing it.
    Tracking entries containing "/" are excluded (treated as no tracking).
    Returns: (has_tracking_dict, missing_tracking_list)
      - has_tracking_dict: {"FBA123": [entries with valid tracking only]}
      - missing_tracking_list: ["FBA456", ...] — FBAs with no valid tracking at all
    """
    has_tracking = {}
    missing_tracking = []
    for fba_id, entries in grouped.items():
        valid = [e for e in entries if e.get("tracking") and "/" not in e["tracking"]]
        if valid:
            has_tracking[fba_id] = valid
        else:
            missing_tracking.append(fba_id)
    return has_tracking, missing_tracking


def _xlrd_cell_str(sheet, row, col) -> str:
    """
    Converts an xlrd cell value to a clean string.
    Numeric cells (e.g. tracking numbers stored as floats) are converted to
    integers first so that 1234567890.0 becomes "1234567890" not "1234567890.0".
    """
    import xlrd
    cell = sheet.cell(row, col)
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        val = cell.value
        # If the float is a whole number, return as integer string.
        return str(int(val)) if val == int(val) else str(val)
    return str(cell.value).strip()


def _detect_xls_sheet_cols(sheet) -> dict:
    """
    Scans the first 3 rows of an xls sheet for a header row containing
    'FBA ID' and 'TRACKING'. Returns a dict:
      {header_row, col_fc, col_fba, col_tracking, col_carrier,
       col_name, col_ctns, col_shipping_way, col_notes}
    Falls back to config-default positions (3, 4, 7, 8) for the core columns if no
    header row is found at all. If a header row IS found but name/ctns/shipping_way
    individually aren't in it, each falls back to its own config-default position
    (1, 5, 6) and logs a warning naming the sheet and field. 'notes' is never
    header-detected — every real sheet carries it in the last physical column
    regardless of that column's header text.
    Note: those name/ctns/shipping_way fallback positions (1, 5, 6) are fixed
    constants defined in this function, NOT read from `config` — unlike the
    .xlsx row-context reader in tracking_status.py's _load_row_context_xlsx,
    which DOES read column_name/column_ctns/column_shipping_way from config.
    This asymmetry has no behavioral effect today (nothing in config.json sets
    those keys), but config values for them would silently be ignored on the
    .xls path.
    """
    name_default, ctns_default, shipping_way_default = 1, 5, 6
    for r in range(min(3, sheet.nrows)):
        vals = [str(sheet.cell(r, c).value).strip().upper() for c in range(sheet.ncols)]
        fba_cols  = [i for i, v in enumerate(vals) if v == "FBA ID"]
        trk_cols  = [i for i, v in enumerate(vals) if "TRACKING" in v]
        dest_cols = [i for i, v in enumerate(vals) if "DESTINATION" in v]
        carr_cols = [i for i, v in enumerate(vals) if v == "CARRIER"]
        name_cols = [i for i, v in enumerate(vals) if "ORDER NO" in v]
        ctns_cols = [i for i, v in enumerate(vals) if "CTNS" in v]
        ship_cols = [i for i, v in enumerate(vals) if "SHIPPING" in v]
        if fba_cols and trk_cols:
            col_trk = trk_cols[0]
            if name_cols:
                col_name = name_cols[0]
            else:
                logger.warning(f"Sheet {sheet.name!r}: could not detect 'name' column from header, falling back to column {name_default}")
                col_name = name_default
            if ctns_cols:
                col_ctns = ctns_cols[0]
            else:
                logger.warning(f"Sheet {sheet.name!r}: could not detect 'ctns' column from header, falling back to column {ctns_default}")
                col_ctns = ctns_default
            if ship_cols:
                col_shipping_way = ship_cols[0]
            else:
                logger.warning(f"Sheet {sheet.name!r}: could not detect 'shipping_way' column from header, falling back to column {shipping_way_default}")
                col_shipping_way = shipping_way_default
            return {
                "header_row": r,
                "col_fc": dest_cols[0] if dest_cols else max(0, fba_cols[0] - 1),
                "col_fba": fba_cols[0],
                "col_tracking": col_trk,
                "col_carrier": carr_cols[0] if carr_cols else col_trk + 1,
                "col_name": col_name,
                "col_ctns": col_ctns,
                "col_shipping_way": col_shipping_way,
                "col_notes": max(0, sheet.ncols - 1),
            }
    return {
        "header_row": 0, "col_fc": 3, "col_fba": 4, "col_tracking": 7, "col_carrier": 8,
        "col_name": name_default, "col_ctns": ctns_default, "col_shipping_way": shipping_way_default,
        "col_notes": max(0, sheet.ncols - 1),
    }


def load_excel_file(file_path: str, config: dict) -> list:
    """
    Loads Excel file rows as dicts using column indices from config.
    For xls files: reads ALL sheets, auto-detects column positions from each sheet's header row.
    For xlsx files: reads ALL sheets using config column indices.
    Default: D=3 (fc_code), E=4 (fba_id), H=7 (tracking_num), I=8 (carrier).
    Skips header rows and rows missing fba_id. Rows with empty tracking_num are included.
    """
    col_fc_cfg = config.get("column_fc_code", 3)
    col_fba_cfg = config.get("column_fba_id", 4)
    col_tracking_cfg = config.get("column_tracking", 7)
    col_carrier_cfg = config.get("column_carrier", 8)
    rows = []

    if detect_excel_engine(file_path) == "xlrd":
        import xlrd
        wb = xlrd.open_workbook(file_path)
        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            cols = _detect_xls_sheet_cols(sheet)
            header_row, col_fc, col_fba = cols["header_row"], cols["col_fc"], cols["col_fba"]
            col_tracking, col_carrier = cols["col_tracking"], cols["col_carrier"]
            for r in range(header_row + 1, sheet.nrows):
                try:
                    fc  = _xlrd_cell_str(sheet, r, col_fc).strip()
                    fba = _xlrd_cell_str(sheet, r, col_fba).strip()
                    trk = _xlrd_cell_str(sheet, r, col_tracking).strip()
                    car = _xlrd_cell_str(sheet, r, col_carrier).strip() if sheet.ncols > col_carrier else ""
                    if fba:
                        rows.append({"fc_code": fc, "fba_id": fba, "tracking_num": trk,
                                     "carrier": car, "row_number": r + 1})
                except IndexError:
                    logger.warning(f"Sheet {sheet.name!r} row {r+1}: IndexError — skipping row")
                    continue
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                try:
                    fc  = str(row[col_fc_cfg] or "").strip()
                    fba = str(row[col_fba_cfg] or "").strip()
                    trk = str(row[col_tracking_cfg] or "").strip()
                    car = str(row[col_carrier_cfg] or "").strip() if len(row) > col_carrier_cfg else ""
                    if fba:
                        rows.append({"fc_code": fc, "fba_id": fba, "tracking_num": trk,
                                     "carrier": car, "row_number": idx + 2})
                except (IndexError, TypeError):
                    logger.warning(f"Sheet {sheet.title!r} row {idx+2}: IndexError/TypeError — skipping row")
                    continue
    return rows


def find_excel_files(input_folder: str) -> list:
    """Returns sorted list of .xls/.xlsx files in input_folder."""
    folder = Path(input_folder)
    if not folder.exists():
        return []
    files = sorted(
        f for pattern in ["*.xls", "*.xlsx"] for f in folder.glob(pattern)
    )
    return [str(f) for f in files]


def parse_and_filter(config: dict) -> dict:
    """
    Top-level: finds Excel files, loads rows, filters US FCs, groups by FBA ID.
    Returns: {"FBA123": [{"tracking": "...", "carrier": "..."}, ...]}
    """
    excel_files = find_excel_files(config["input_folder"])
    if not excel_files:
        logger.warning(f"No Excel files found in {config['input_folder']}")
        return {}
    if len(excel_files) > 1:
        logger.warning(f"Multiple Excel files found — processing all: {excel_files}")

    us_prefixes = load_us_fc_prefixes(config.get("us_fc_codes_file", "us_fc_codes.txt"))
    if not us_prefixes:
        logger.warning("No US FC prefixes loaded — check us_fc_codes.txt")

    all_us_rows = []
    for file_path in excel_files:
        logger.info(f"Reading: {file_path}")
        all_rows = load_excel_file(file_path, config)
        us_rows = [r for r in all_rows if is_us_fc(r["fc_code"], us_prefixes)]
        logger.info(f"  {len(us_rows)} US rows (skipped {len(all_rows) - len(us_rows)} non-US)")
        all_us_rows.extend(us_rows)

    return group_by_fba_id(all_us_rows)


def parse_and_filter_by_region_full(config: dict) -> tuple:
    """
    Same as parse_and_filter_by_region, but also returns the raw rows that
    didn't match any region's FC code prefixes.
    Returns: (region_dict, unmatched_rows)
      - region_dict: {"US": {"FBA123": [...]}, ...} — identical shape to parse_and_filter_by_region()
      - unmatched_rows: [{"fc_code": ..., "fba_id": ..., "tracking_num": ..., "carrier": ..., "row_number": ...}, ...]
    """
    regions = config.get("regions", [])
    if not regions:
        logger.warning("No 'regions' key in config — falling back to US-only parse_and_filter()")
        return {"US": parse_and_filter(config)}, []

    excel_files = find_excel_files(config["input_folder"])
    if not excel_files:
        logger.warning(f"No Excel files found in {config['input_folder']}")
        return {r["name"]: {} for r in regions}, []

    all_rows = []
    for file_path in excel_files:
        logger.info(f"Reading: {file_path}")
        rows = load_excel_file(file_path, config)
        all_rows.extend(rows)
    logger.info(f"Loaded {len(all_rows)} total rows across {len(excel_files)} file(s)")

    result = {}
    matched_row_ids = set()
    for region in regions:
        name = region["name"]
        fc_file = region.get("fc_codes_file", "")
        prefixes = load_fc_prefixes(fc_file)
        if not prefixes:
            logger.warning(f"[{name}] No FC prefixes loaded from {fc_file!r}")
        region_rows = [r for r in all_rows if is_region_fc(r["fc_code"], prefixes)]
        logger.info(f"[{name}] {len(region_rows)} row(s) matched")
        matched_row_ids.update(id(r) for r in region_rows)
        result[name] = group_by_fba_id(region_rows)

    unmatched_rows = [r for r in all_rows if id(r) not in matched_row_ids]
    return result, unmatched_rows


def parse_and_filter_by_region(config: dict) -> dict:
    """
    Finds Excel files, loads all rows, then splits by region using each region's FC codes file.
    Returns: {"US": {"FBA123": [...]}, "CA": {"FBA456": [...]}, ...}
    Each region only contains FBA IDs whose FC code matches that region's prefixes.
    """
    region_dict, _ = parse_and_filter_by_region_full(config)
    return region_dict
