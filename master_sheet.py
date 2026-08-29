# master_sheet.py
"""
Persistent workbook consolidating carrier-tracking status and Send-to-Amazon
workflow/delivery-window state, one row per FBA ID. Rewritten in full from an
in-memory dict on every save (same convention as tracking_status.py's status
cache), so updating a shipment's row is just: load, mutate the dict entry for
that FBA ID, save.
"""
from pathlib import Path

MASTER_SHEET_PATH_DEFAULT = "logs/shipment_tracking_master.xlsx"

MASTER_SHEET_COLUMNS = [
    "Tracking Status", "Delivery Date Status", "Tracking Number", "Carrier",
    "FBA ID", "Shipment Name", "Destination", "Ctns", "Shipping Way",
    "Notes (source)", "Label Created Date", "Expected Delivery Date",
    "Current Status", "Last Checked", "Region", "Workflow ID",
    "Delivery Window Start", "Delivery Window End", "Delivery Window Last Checked",
]

# Maps in-memory dict keys to their column position/header above, in column order.
_FIELD_ORDER = [
    ("tracking_status", "Tracking Status"),
    ("delivery_date_status", "Delivery Date Status"),
    ("tracking", "Tracking Number"),
    ("carrier", "Carrier"),
    ("fba_id", "FBA ID"),
    ("name", "Shipment Name"),
    ("destination", "Destination"),
    ("ctns", "Ctns"),
    ("shipping_way", "Shipping Way"),
    ("notes", "Notes (source)"),
    ("label_created_date", "Label Created Date"),
    ("expected_delivery_date", "Expected Delivery Date"),
    ("status", "Current Status"),
    ("last_checked", "Last Checked"),
    ("region", "Region"),
    ("workflow_id", "Workflow ID"),
    ("delivery_window_start", "Delivery Window Start"),
    ("delivery_window_end", "Delivery Window End"),
    ("delivery_window_last_checked", "Delivery Window Last Checked"),
]


def load_master_sheet(path: str) -> dict:
    """Reads the persistent master workbook into {fba_id: {field: value}}."""
    if not Path(path).exists():
        return {}
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    sheet = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[4]:  # FBA ID column
            continue
        entry = {}
        for i, (key, _header) in enumerate(_FIELD_ORDER):
            value = row[i] if i < len(row) else None
            entry[key] = value if value is not None else ""
        sheet[str(entry["fba_id"]).strip()] = entry
    return sheet


# Fields sourced fresh from the input sheet on every populate — refreshed for
# shipments already in the master sheet in case the supplier's data changed.
_SOURCE_FIELDS = ["tracking", "carrier", "name", "destination", "ctns", "shipping_way", "notes", "region"]

# Fields owned by later processing (carrier checks, workflow-ID discovery,
# window sync) — never touched by populate_from_input for a shipment that's
# already in the sheet.
_STATUS_FIELDS = [
    "tracking_status", "delivery_date_status", "label_created_date",
    "expected_delivery_date", "status", "last_checked", "workflow_id",
    "delivery_window_start", "delivery_window_end", "delivery_window_last_checked",
]


def populate_from_input(config: dict, master_sheet: dict) -> dict:
    """
    Seeds/refreshes the master sheet dict from the input Excel file, one row per
    FBA ID (shipments with blank tracking are skipped, matching build_check_list).
    New FBA IDs get a fresh row with tracking_status/delivery_date_status
    "pending" and no workflow_id yet. FBA IDs already present in master_sheet
    keep their status/workflow fields untouched -- only their source fields
    (tracking, carrier, name, destination, ctns, shipping_way, notes, region)
    are refreshed, in case the supplier's sheet changed since the last run.
    Returns a new dict; does not mutate master_sheet in place.
    """
    from tracking_status import build_check_list

    result = {fba_id: dict(entry) for fba_id, entry in master_sheet.items()}
    for entry in build_check_list(config):
        fba_id = entry["fba_id"]
        source = {field: entry.get(field, "") for field in _SOURCE_FIELDS}
        if fba_id in result:
            result[fba_id].update(source)
        else:
            row = dict(source)
            row["fba_id"] = fba_id
            row["tracking_status"] = "pending"
            row["delivery_date_status"] = "pending"
            row["workflow_id"] = ""
            row["label_created_date"] = ""
            row["expected_delivery_date"] = ""
            row["status"] = ""
            row["last_checked"] = ""
            row["delivery_window_start"] = ""
            row["delivery_window_end"] = ""
            row["delivery_window_last_checked"] = ""
            result[fba_id] = row
    return result


def sync_delivered_status(sheet: dict, tracking_cache: dict) -> dict:
    """
    For every row not already marked "Delivered", checks whether it actually
    is now -- either the source notes already say so (the supplier's own
    "delivered on ..." note), or logs/tracking_status.xlsx's cache (matched by
    tracking number) reports status == "Delivered" from a live carrier check.
    If so, both tracking_status and delivery_date_status are set to
    "Delivered" -- there's nothing left to track on either front once a
    shipment has actually arrived. Returns a new dict; does not mutate sheet.
    """
    from tracking_status import _notes_indicate_delivered

    result = {fba_id: dict(entry) for fba_id, entry in sheet.items()}
    for entry in result.values():
        if entry.get("tracking_status") == "Delivered":
            continue
        delivered = _notes_indicate_delivered(entry.get("notes", ""))
        if not delivered:
            cached = tracking_cache.get(str(entry.get("tracking", "")).strip())
            delivered = bool(cached) and cached.get("status") == "Delivered"
        if delivered:
            entry["tracking_status"] = "Delivered"
            entry["delivery_date_status"] = "Delivered"
    return result


def run_update_master_sheet(config: dict) -> dict:
    """
    Loads the persistent master sheet, populates/refreshes it from the input
    Excel file, syncs delivered status from notes/the tracking-status cache,
    and saves it back. Used both as the standalone --update-master-sheet CLI
    mode and as a step in the main pipeline. Returns {"path", "total", "new"}.
    """
    from tracking_status import load_status_cache, STATUS_CACHE_PATH_DEFAULT

    path = config.get("master_sheet_path", MASTER_SHEET_PATH_DEFAULT)
    existing = load_master_sheet(path)
    before = len(existing)
    sheet = populate_from_input(config, existing)

    tracking_cache_path = config.get("tracking_status_cache", STATUS_CACHE_PATH_DEFAULT)
    tracking_cache = load_status_cache(tracking_cache_path)
    sheet = sync_delivered_status(sheet, tracking_cache)

    save_master_sheet(path, sheet)
    return {"path": path, "total": len(sheet), "new": len(sheet) - before}


def format_update_master_sheet_summary(result: dict) -> str:
    lines = [
        "=" * 60,
        "MASTER SHEET UPDATE SUMMARY",
        "=" * 60,
        f"Total shipments in sheet: {result['total']}",
        f"New shipments added:      {result['new']}",
        f"Saved to:                 {result['path']}",
        "=" * 60,
    ]
    return "\n".join(lines)


def save_master_sheet(path: str, sheet: dict) -> None:
    """Rewrites the whole master workbook from the in-memory sheet dict, sorted by FBA ID."""
    import openpyxl

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shipment Tracking Master"
    ws.append(MASTER_SHEET_COLUMNS)
    for fba_id in sorted(sheet.keys()):
        entry = sheet[fba_id]
        ws.append([entry.get(key, "") for key, _header in _FIELD_ORDER])
    wb.save(path)
