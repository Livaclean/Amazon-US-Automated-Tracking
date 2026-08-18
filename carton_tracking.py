# carton_tracking.py
"""
Parses the freeform "carton tracking" blob some source sheets carry in a
trailing column: newline-separated entries pairing a physical carrier
tracking number to one specific carton of one specific destination FBA
shipment, e.g. "1ZK6B4420338604208-FBA19L9DHD1SU000001" (tracking number,
FBA ID, U-prefixed 6-digit carton sequence). When present and well-formed,
this gives ground-truth per-carton tracking assignment — no carrier-website
scrape or Amazon-slot-count guessing needed.
"""
import re
import logging

from fetch_sub_tracking import UPS_PATTERN, FEDEX_PATTERN

logger = logging.getLogger(__name__)

# FBA IDs in this blob are consistently "FBA" + 9 alphanumeric chars (12
# total), immediately followed by "U" + a 6-digit carton sequence with no
# separator — e.g. "FBA19L9DHD1SU000001".
_FBA_SEQ_RE = re.compile(r"(FBA[A-Z0-9]{9})U(\d{6})", re.IGNORECASE)

_TRACKING_RES = [UPS_PATTERN, FEDEX_PATTERN]


def _find_tracking_matches(line: str) -> list:
    """Returns all UPS/FedEx tracking-number matches found in line."""
    matches = []
    for pattern in _TRACKING_RES:
        matches.extend(pattern.findall(line))
    return matches


def parse_carton_tracking_cell(text, warn: bool = True) -> list:
    """
    Parses a carton-tracking blob cell into an ordered list of
    {"fba_id": str, "seq": int, "tracking": str} dicts, one per line.

    Returns:
      [] if text is blank (no data — not an error).
      None if any line doesn't contain exactly one tracking number paired
        with exactly one FBA+sequence token — ambiguous data is never
        guessed; caller should fall back to the carrier-website scrape.

    warn=False suppresses the per-line mismatch log (used while probing
    candidate columns, where most cells are expected not to match at all).
    """
    if not text or not str(text).strip():
        return []

    entries = []
    for raw_line in str(text).splitlines():
        line = raw_line.strip()
        if not line:
            continue

        fba_matches = _FBA_SEQ_RE.findall(line)
        tracking_matches = _find_tracking_matches(line.upper())

        if len(fba_matches) != 1 or len(tracking_matches) != 1:
            if warn:
                logger.warning(
                    f"Unparseable carton-tracking line (expected 1 FBA+seq and "
                    f"1 tracking, got {len(fba_matches)} and {len(tracking_matches)}): {line!r}"
                )
            return None

        fba_id, seq = fba_matches[0]
        entries.append({
            "fba_id": fba_id.upper(),
            "seq": int(seq),
            "tracking": tracking_matches[0].upper(),
        })

    return entries


def _split_fba_id_parts(fba_id: str) -> list:
    """Mirrors parse_excel.group_by_fba_id()'s '/'-combined FBA ID splitting."""
    if "/" not in fba_id:
        return [fba_id]
    return [p.strip() for p in fba_id.split("/") if p.strip()]


def _register_matches(result: dict, fba_id: str, row_number: int, entries: list) -> None:
    """For each '/'-split part of fba_id, finds this row's own carton entries
    within `entries` (filtered by matching FBA ID, ordered by carton sequence)
    and stores them in result under (part, row_number)."""
    for part in _split_fba_id_parts(fba_id):
        matched = sorted(
            (e for e in entries if e["fba_id"] == part.upper()),
            key=lambda e: e["seq"],
        )
        if matched:
            result[(part, row_number)] = [e["tracking"] for e in matched]


def build_carton_tracking_map(file_paths: list, config: dict) -> dict:
    """
    Scans source Excel files for a carton-tracking blob column and returns
    {(fba_id, row_number): [tracking_1, tracking_2, ...]} — tracking numbers
    ordered by carton sequence, filtered to just the entries whose embedded
    FBA ID matches that row's own FBA ID (a shared blob can list cartons
    destined for OTHER FBA shipments in the same freight batch — see
    parse_carton_tracking_cell's docstring). Rows whose blob cell fails to
    parse (ambiguous/malformed) are simply absent from the result — logged by
    parse_carton_tracking_cell, not raised — so callers naturally fall back
    to the carrier-website scrape for that row.
    Keyed the same way parse_excel.load_excel_file() computes row_number
    (xls: r + 1) so callers can look up the exact physical row, matching the
    convention tracking_status.load_row_context() already uses.
    Only .xls files are content-scanned today (matches the real source
    format); .xlsx files are only scanned if config sets "column_cartons".
    """
    from parse_excel import detect_excel_engine, _detect_xls_sheet_cols, _xlrd_cell_str

    result = {}
    for file_path in file_paths:
        if detect_excel_engine(file_path) == "xlrd":
            import xlrd
            wb = xlrd.open_workbook(file_path)
            for sheet_idx in range(wb.nsheets):
                sheet = wb.sheet_by_index(sheet_idx)
                cols = _detect_xls_sheet_cols(sheet)
                col_cartons = cols["col_cartons"]
                if col_cartons is None:
                    continue
                for r in range(cols["header_row"] + 1, sheet.nrows):
                    fba_id = _xlrd_cell_str(sheet, r, cols["col_fba"]).strip()
                    if not fba_id:
                        continue
                    entries = parse_carton_tracking_cell(_xlrd_cell_str(sheet, r, col_cartons))
                    if not entries:
                        continue
                    _register_matches(result, fba_id, r + 1, entries)
        else:
            col_cartons_cfg = config.get("column_cartons")
            if col_cartons_cfg is None:
                continue
            from openpyxl import load_workbook
            col_fba_cfg = config.get("column_fba_id", 4)
            wb = load_workbook(file_path, read_only=True, data_only=True)
            for sheet in wb.worksheets:
                for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                    if col_fba_cfg >= len(row):
                        continue
                    fba_id = str(row[col_fba_cfg] or "").strip()
                    if not fba_id or col_cartons_cfg >= len(row):
                        continue
                    entries = parse_carton_tracking_cell(str(row[col_cartons_cfg] or ""))
                    if not entries:
                        continue
                    _register_matches(result, fba_id, idx + 2, entries)
    return result


def merge_carton_tracking_for_fba(carton_map: dict, fba_id: str, entries: list) -> list:
    """
    Merges column-L carton-tracking IDs across all physical rows an FBA spans,
    deduplicated while preserving order. carton_map is keyed (fba_id, row_number)
    -> [tracking, ...], as returned by build_carton_tracking_map(). Shared by
    run.py's main upload flow and verify_tracking.py's re-upload flow.
    Empty list means no column-L data was found for this FBA at all.
    """
    merged = []
    for entry in entries:
        merged.extend(carton_map.get((fba_id, entry.get("row_number")), []))
    seen = set()
    result = []
    for t in merged:
        if t not in seen:
            seen.add(t)
            result.append(t)
    return result


def expected_carton_count_for_fba(row_ctx: dict, fba_id: str, entries: list):
    """
    Sums the "NO OF CTNS" value across all physical rows an FBA spans, using
    row_ctx keyed (fba_id, row_number) -> {"ctns": ..., ...} (same shape as
    tracking_status.load_row_context()). Returns None if no row context (or
    no parseable ctns value) was found for this FBA at all.
    """
    total = 0
    found = False
    for entry in entries:
        ctx = row_ctx.get((fba_id, entry.get("row_number")))
        if not ctx:
            continue
        try:
            total += int(float(ctx.get("ctns", "")))
            found = True
        except (ValueError, TypeError):
            continue
    return total if found else None


def detect_carton_tracking_column(sheet, header_row: int, exclude_cols: set) -> int:
    """
    Scans a sheet's data rows (content-based, not positional) for the column
    holding carton-tracking blobs. Returns the column index, or None if no
    column in the sheet contains any successfully-parsed, non-empty entry.
    exclude_cols are column indices already assigned other meaning (fba,
    tracking, carrier, etc.) and are never considered.
    """
    for col in range(sheet.ncols):
        if col in exclude_cols:
            continue
        for r in range(header_row + 1, sheet.nrows):
            value = sheet.cell(r, col).value
            parsed = parse_carton_tracking_cell(value, warn=False)
            if parsed:
                return col
    return None
