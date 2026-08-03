# tracking_status.py
import logging
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

STATUS_CACHE_PATH_DEFAULT = "logs/tracking_status.xlsx"

# Status/date extraction is regex-based text scanning over the carrier page's
# visible text. Like the selectors in upload_tracking.py's SELECTORS dict, these
# phrases are a starting point — run against real tracking pages and refine if a
# carrier's wording doesn't match (see upload_tracking.py's --discover convention).
_DATE_PATTERN = re.compile(
    r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},?\s+\d{4}\b"
    # UPS's "Delivered" banner shows a bare weekday + month + day with no year,
    # e.g. "Friday, July 17" — must come before the dated-month branch's word chars
    # would otherwise be consumed without matching (no \d{4} present here).
    r"|\b(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)[a-z]*,?\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2}\b"
    r"|\b\d{1,2}/\d{1,2}/\d{2,4}\b"
    r"|\b\d{4}-\d{2}-\d{2}\b",
    re.IGNORECASE,
)

_STATUS_KEYWORDS = [
    ("Invalid tracking number", "Invalid"),
    ("Delivered", "Delivered"),
    ("Out for Delivery", "Out for Delivery"),
    ("Exception", "Exception"),
    ("In Transit", "In Transit"),
    ("On the Way", "In Transit"),
    ("Picked Up", "Picked Up"),
    ("Shipment Information Sent", "Label Created"),
    ("Shipper Created a Label", "Label Created"),
    ("Label Created", "Label Created"),
    ("Order Processed", "Label Created"),
]

_LABEL_CREATED_ANCHORS = [
    "Label Created", "Shipment Information Sent", "Shipper Created a Label", "Order Processed",
    "Shipped / Billed On",
]
_EXPECTED_DELIVERY_ANCHORS = [
    "Scheduled Delivery", "Estimated Delivery", "Expected Delivery", "Delivery Date",
    "Delivered",  # already-delivered UPS pages show a bare date right after this banner
]

# Some carrier pages (UPS in particular) list every possible milestone as plain
# text regardless of which one is current, e.g. "Out for Delivery" appears on
# every page as a future step even when the real status is "Label Created".
# Each milestone is followed by a literal "active"/"inactive" line — only the
# one marked "active" reflects the real current status.
_MILESTONE_STATUS_MAP = {
    "Delivered": "Delivered",
    "Delivery": "Delivered",
    "Out for Delivery": "Out for Delivery",
    "On the Way": "In Transit",
    "In Transit": "In Transit",
    "We Have Your Package": "Picked Up",
    "Picked Up": "Picked Up",
    "Label Created": "Label Created",
    "Shipment Information Sent": "Label Created",
}

_UNKNOWN_STATUS = {"status": "Unknown", "label_created_date": None, "expected_delivery_date": None}


def _extract_active_milestone_status(text: str):
    """Finds the milestone whose next non-blank line is exactly 'active' (not 'inactive')."""
    if not text:
        return None
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for i, line in enumerate(lines[:-1]):
        if line in _MILESTONE_STATUS_MAP and lines[i + 1].strip().lower() == "active":
            return _MILESTONE_STATUS_MAP[line]
    return None


def _extract_fedex_from_line_status(text: str):
    """
    FedEx always lists every milestone (WE HAVE YOUR PACKAGE / ON THE WAY / OUT FOR
    DELIVERY) as plain text with no active/inactive markers, so those can't be used
    to find the current step like on UPS. The real current status instead appears
    once, as the line directly under "FROM" (e.g. "Label created").
    """
    if not text:
        return None
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for i, line in enumerate(lines[:-1]):
        if line.upper() == "FROM":
            normalized = _extract_status_from_text(lines[i + 1])
            if normalized != "Unknown":
                return normalized
    return None


def _extract_status_from_text(text: str) -> str:
    if not text:
        return "Unknown"
    lower = text.lower()
    for phrase, normalized in _STATUS_KEYWORDS:
        if phrase.lower() in lower:
            return normalized
    return "Unknown"


def _nearby_date(text: str, anchor_phrases: list, window: int = 200):
    """Finds the first date-like string within `window` chars after any anchor phrase."""
    if not text:
        return None
    lower = text.lower()
    for phrase in anchor_phrases:
        idx = lower.find(phrase.lower())
        if idx == -1:
            continue
        segment = text[idx: idx + window]
        m = _DATE_PATTERN.search(segment)
        if m:
            return m.group(0)
    return None


def _extract_status_and_dates_from_text(text: str) -> dict:
    status = (
        _extract_active_milestone_status(text)
        or _extract_fedex_from_line_status(text)
        or _extract_status_from_text(text)
    )
    return {
        "status": status,
        "label_created_date": _nearby_date(text, _LABEL_CREATED_ANCHORS),
        "expected_delivery_date": _nearby_date(text, _EXPECTED_DELIVERY_ANCHORS),
    }


_DATE_IN_NOTES_RE = re.compile(r"([0-9]{1,4}[./-][0-9]{1,2}(?:[./-][0-9]{1,4})?)")


def _notes_indicate_delivered(notes) -> bool:
    """True if the source sheet's freeform notes column (K) already says 'delivered'."""
    return bool(notes) and "delivered" in str(notes).lower()


def _extract_date_from_notes(notes):
    """Pulls the first date-like token out of a notes string, e.g. 'delivered on 2026.02.24'."""
    if not notes:
        return None
    m = _DATE_IN_NOTES_RE.search(str(notes))
    return m.group(1) if m else None


def seed_delivered_from_notes(grouped: dict, cache: dict) -> dict:
    """
    If the source sheet's notes column already records a shipment as delivered
    (suppliers hand-write things like "delivered on 2026.02.24"), trust that
    instead of hitting the carrier for it — it's the supplier's own record and
    saves a scrape. Returns a new cache dict; does not mutate the input.
    """
    updated = dict(cache)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    for tracking, entries in grouped.items():
        if updated.get(tracking, {}).get("status") == "Delivered":
            continue
        delivered_entry = next(
            (e for e in entries if _notes_indicate_delivered(e.get("notes", ""))), None
        )
        if not delivered_entry:
            continue
        updated[tracking] = {
            "carrier": delivered_entry.get("carrier", ""),
            "fba_id": delivered_entry.get("fba_id", ""),
            "name": delivered_entry.get("name", ""),
            "destination": delivered_entry.get("destination", ""),
            "ctns": delivered_entry.get("ctns", ""),
            "shipping_way": delivered_entry.get("shipping_way", ""),
            "notes": delivered_entry.get("notes", ""),
            "label_created_date": updated.get(tracking, {}).get("label_created_date"),
            "expected_delivery_date": _extract_date_from_notes(delivered_entry.get("notes", "")),
            "status": "Delivered",
            "last_checked": now,
            "region": delivered_entry.get("region", ""),
        }
    return updated


def _check_ups_status(page, tracking: str, logs_folder: str = None) -> dict:
    from fetch_sub_tracking import UPS_TRACK_URL, _handle_captcha

    url = UPS_TRACK_URL.format(tracking=tracking)
    logger.info(f"  Loading: {url}")
    try:
        page.goto(url, timeout=30000)
        page.wait_for_load_state("load", timeout=30000)
        page.wait_for_timeout(6000)
    except Exception as e:
        logger.error(f"  UPS status check failed to load {tracking}: {e}")
        return dict(_UNKNOWN_STATUS)

    _handle_captcha(page)

    try:
        text = page.inner_text("body")
    except Exception as e:
        logger.error(f"  UPS page text extraction failed for {tracking}: {e}")
        return dict(_UNKNOWN_STATUS)

    return _extract_status_and_dates_from_text(text)


_FEDEX_KEY_STATUS_MAP = {
    "label created": "Label Created",
    "in transit": "In Transit",
    "on the way": "In Transit",
    "picked up": "Picked Up",
    "out for delivery": "Out for Delivery",
    "delivered": "Delivered",
}


def _fedex_status_from_api(data: dict) -> dict:
    """
    Extracts status/dates from FedEx's real backend API (api.fedex.com/track/v2/shipments),
    captured live via network interception. Far more reliable than scraping the rendered
    page: keyStatus is the actual current status (no milestone-bar ambiguity), and
    scanEventList carries the "Shipment information sent to FedEx" event with the real
    label-created date — a date the rendered page doesn't display at all pre-pickup.
    """
    packages = (data.get("output", {}) or {}).get("packages", [])
    if not packages:
        return dict(_UNKNOWN_STATUS)
    pkg = packages[0]

    key_status = str(pkg.get("keyStatus", "") or "").strip().lower()
    status = _FEDEX_KEY_STATUS_MAP.get(key_status) or _extract_status_from_text(key_status)

    label_created_date = None
    for event in pkg.get("scanEventList", []) or []:
        if "shipment information sent" in str(event.get("status", "")).lower():
            label_created_date = event.get("date")
            break
    if not label_created_date:
        ship_dt = pkg.get("displayShipDt", "") or ""
        if ship_dt and "will be updated" not in ship_dt.lower():
            label_created_date = ship_dt

    expected_delivery_date = None
    act_delivery = pkg.get("displayActDeliveryDt", "") or ""
    est_delivery = pkg.get("displayEstDeliveryDt", "") or ""
    if act_delivery:
        expected_delivery_date = act_delivery
    elif est_delivery and "will be updated" not in est_delivery.lower():
        expected_delivery_date = est_delivery

    return {
        "status": status,
        "label_created_date": label_created_date,
        "expected_delivery_date": expected_delivery_date,
    }


def _check_fedex_status(page, tracking: str, logs_folder: str = None) -> dict:
    """
    Loads the plain trknbr tracking URL and lets FedEx's own client-side redirect
    resolve the internal trkqual piece-qualifier itself. Guessing the qualifier
    (as fetch_sub_tracking.py's sub-tracking fetcher does) was tried first here but
    landed on a stuck skeleton "Loading..." page when the guess was wrong — letting
    FedEx resolve it itself is both simpler and more reliable for this purpose.

    Primary path: intercepts the track/v2/shipments JSON API the page itself calls
    (same technique as fetch_dhl_sub_tracking's utapi interception) for clean
    structured data, including the label-created date the rendered page never shows.
    Falls back to page-text scraping if interception fails.
    """
    from fetch_sub_tracking import _handle_captcha

    url = f"https://www.fedex.com/fedextrack/?trknbr={tracking}"
    logger.info(f"  Loading: {url}")

    try:
        with page.expect_response(
            lambda r: r.url.rstrip("/").endswith("/track/v2/shipments") and r.status == 200,
            timeout=25000,
        ) as resp_ctx:
            page.goto(url, timeout=30000)
        result = _fedex_status_from_api(resp_ctx.value.json())
        if result["status"] != "Unknown":
            return result
        logger.debug(f"  FedEx API gave no usable status for {tracking} — falling back to page text")
    except Exception as e:
        logger.debug(f"  FedEx API interception failed for {tracking}: {e} — falling back to page text")

    try:
        page.wait_for_load_state("domcontentloaded", timeout=15000)
        page.wait_for_timeout(10000)
    except Exception:
        pass

    _handle_captcha(page)

    try:
        text = page.inner_text("body")
    except Exception as e:
        logger.error(f"  FedEx page text extraction failed for {tracking}: {e}")
        return dict(_UNKNOWN_STATUS)

    if "can't find" in text.lower():
        return dict(_UNKNOWN_STATUS)

    return _extract_status_and_dates_from_text(text)


_DHL_STATUS_CODE_MAP = {
    "pre-transit": "Label Created",
    "transit": "In Transit",
    "delivered": "Delivered",
    "failure": "Exception",
}


def _dhl_status_from_api(data: dict) -> dict:
    """
    Extracts status/dates from DHL's utapi JSON (same endpoint fetch_dhl_sub_tracking
    already intercepts for pieceIds). details.shipmentActivationDate is the real
    label-created date — a keyword scan over `events` is unreliable since many transit
    event descriptions (e.g. "Shipment is out with courier for delivery") contain the
    word "shipment" too and get matched instead of the actual label/pickup event.
    """
    shipments = data.get("shipments", [])
    if not shipments:
        return dict(_UNKNOWN_STATUS)
    shipment = shipments[0]

    status_obj = shipment.get("status", {}) or {}
    status_code = str(status_obj.get("statusCode", "") or "").strip().lower()
    status_text = status_obj.get("description") or status_code or ""
    status = _DHL_STATUS_CODE_MAP.get(status_code) or (
        _extract_status_from_text(status_text) if status_text else "Unknown"
    )

    label_created_date = (shipment.get("details", {}) or {}).get("shipmentActivationDate")

    if status == "Delivered":
        expected_delivery_date = status_obj.get("timestamp")
    else:
        expected_delivery_date = shipment.get("estimatedTimeOfDelivery") or shipment.get("estimatedDeliveryDate")

    return {
        "status": status,
        "label_created_date": label_created_date,
        "expected_delivery_date": expected_delivery_date,
    }


def _check_dhl_status(page, tracking: str, logs_folder: str = None) -> dict:
    from fetch_sub_tracking import DHL_TRACK_URL, _handle_captcha

    url = DHL_TRACK_URL.format(tracking=tracking)
    logger.info(f"  Loading: {url}")

    try:
        with page.expect_response(
            lambda r: f"utapi?trackingNumber={tracking}" in r.url and r.status == 200,
            timeout=20000,
        ) as resp_ctx:
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
        result = _dhl_status_from_api(resp_ctx.value.json())
        if result["status"] != "Unknown":
            return result
        logger.debug(f"  DHL utapi gave no usable status for {tracking} — falling back to page text")
    except Exception as e:
        logger.debug(f"  DHL utapi interception failed for {tracking}: {e} — falling back to page text")

    try:
        page.wait_for_load_state("load", timeout=15000)
        page.wait_for_timeout(3000)
    except Exception:
        pass

    _handle_captcha(page)

    try:
        text = page.inner_text("body")
    except Exception as e:
        logger.error(f"  DHL page text extraction failed for {tracking}: {e}")
        return dict(_UNKNOWN_STATUS)

    return _extract_status_and_dates_from_text(text)


def check_tracking_status(page, tracking: str, carrier: str, logs_folder: str = None) -> dict:
    """
    Visits the carrier's tracking page for `tracking` and returns
    {"status": str, "label_created_date": str|None, "expected_delivery_date": str|None}.
    Never raises — returns an "Unknown" result on any failure so one bad tracking
    number doesn't stop a run.
    """
    from fetch_sub_tracking import normalize_carrier, _detect_carrier_from_tracking

    normalized = normalize_carrier(carrier)
    if normalized == "unknown":
        detected = _detect_carrier_from_tracking(tracking)
        if detected != "unknown":
            normalized = detected

    try:
        if normalized == "ups":
            return _check_ups_status(page, tracking, logs_folder)
        elif normalized == "fedex":
            return _check_fedex_status(page, tracking, logs_folder)
        elif normalized == "dhl":
            return _check_dhl_status(page, tracking, logs_folder)
        else:
            logger.warning(f"  Unknown carrier '{carrier}' for tracking {tracking} — skipping status check")
            return dict(_UNKNOWN_STATUS)
    except Exception as e:
        logger.error(f"  check_tracking_status failed for {tracking} ({carrier}): {e}")
        return dict(_UNKNOWN_STATUS)

STATUS_CACHE_COLUMNS = [
    "Tracking Number", "Carrier", "FBA ID", "Shipment Name", "Destination",
    "Ctns", "Shipping Way", "Notes (source)", "Label Created Date",
    "Expected Delivery Date", "Current Status", "Last Checked", "Region",
]

# Maps cache dict keys (as used in-memory) to their column position/header above.
_CACHE_FIELD_ORDER = [
    ("carrier", "Carrier"),
    ("fba_id", "FBA ID"),
    ("name", "Shipment Name"),
    ("destination", "Destination"),
    ("ctns", "Ctns"),
    ("shipping_way", "Shipping Way"),
    ("notes", "Notes (source)"),
    ("label_created_date", "Label Created Date"),
    ("expected_delivery_date", "Expected Delivery Date"),
    ("status", "Current Status"),
    ("last_checked", "Last Checked"),
    ("region", "Region"),
]


@dataclass
class CheckTrackingResult:
    checked: int
    skipped_delivered: int
    errors: int
    updated_at: str


def load_row_context(file_path: str, config: dict) -> dict:
    """
    Reads descriptive columns (name, destination, ctns, shipping_way, notes) from an
    xlsx file, keyed by 1-indexed row_number, matching parse_excel.load_excel_file's
    row_number convention (min_row=2, row_number = idx + 2).
    """
    from openpyxl import load_workbook

    col_fc = config.get("column_fc_code", 3)
    col_name = config.get("column_name", 1)
    col_ctns = config.get("column_ctns", 5)
    col_shipping_way = config.get("column_shipping_way", 6)
    col_notes = config.get("column_notes", 10)

    context = {}
    wb = load_workbook(file_path, read_only=True, data_only=True)
    for sheet in wb.worksheets:
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            row_number = idx + 2
            try:
                context[row_number] = {
                    "name": str(row[col_name] or "").strip(),
                    "destination": str(row[col_fc] or "").strip(),
                    "ctns": row[col_ctns] if col_ctns < len(row) else "",
                    "shipping_way": str(row[col_shipping_way] or "").strip(),
                    "notes": str(row[col_notes] or "").strip() if col_notes < len(row) else "",
                }
            except (IndexError, TypeError):
                logger.warning(f"Sheet {sheet.title!r} row {row_number}: IndexError/TypeError — skipping context")
                continue
    return context


def group_by_tracking(entries: list) -> dict:
    """
    Groups shipment entries by tracking number so the carrier is only checked once
    per unique tracking number, even when multiple FBA rows share it.
    Entries with blank tracking are skipped.
    """
    grouped = {}
    for entry in entries:
        tracking = str(entry.get("tracking") or "").strip()
        if not tracking:
            continue
        grouped.setdefault(tracking, []).append(entry)
    return grouped


def build_check_list(config: dict) -> list:
    """
    Combines parse_and_filter_by_region() shipment data with descriptive row context
    (name, destination, ctns, shipping_way, notes) into one flat list of dicts:
    {region, fba_id, tracking, carrier, row_number, name, destination, ctns, shipping_way, notes}.
    Entries with blank tracking are skipped.
    """
    from parse_excel import parse_and_filter_by_region, find_excel_files

    by_region = parse_and_filter_by_region(config)

    context = {}
    for file_path in find_excel_files(config["input_folder"]):
        context.update(load_row_context(file_path, config))

    result = []
    for region_name, fba_shipments in by_region.items():
        for fba_id, tracking_entries in fba_shipments.items():
            for entry in tracking_entries:
                tracking = str(entry.get("tracking") or "").strip()
                if not tracking:
                    continue
                row_ctx = context.get(entry.get("row_number"), {})
                result.append({
                    "region": region_name,
                    "fba_id": fba_id,
                    "tracking": tracking,
                    "carrier": entry.get("carrier", ""),
                    "row_number": entry.get("row_number"),
                    "name": row_ctx.get("name", ""),
                    "destination": row_ctx.get("destination", ""),
                    "ctns": row_ctx.get("ctns", ""),
                    "shipping_way": row_ctx.get("shipping_way", ""),
                    "notes": row_ctx.get("notes", ""),
                })
    return result


def load_status_cache(path: str) -> dict:
    """Reads the persistent tracking-status workbook into {tracking_number: {field: value}}."""
    if not Path(path).exists():
        return {}
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    cache = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        tracking = str(row[0]).strip()
        entry = {}
        for i, (key, _header) in enumerate(_CACHE_FIELD_ORDER, start=1):
            entry[key] = row[i] if i < len(row) else ""
        cache[tracking] = entry
    return cache


def save_status_cache(path: str, cache: dict) -> None:
    """Rewrites the whole tracking-status workbook from the in-memory cache dict."""
    import openpyxl

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracking Status"
    ws.append(STATUS_CACHE_COLUMNS)
    for tracking in sorted(cache.keys()):
        entry = cache[tracking]
        row = [tracking] + [entry.get(key, "") for key, _header in _CACHE_FIELD_ORDER]
        ws.append(row)
    wb.save(path)


def partition_by_delivered(grouped: dict, cache: dict) -> tuple:
    """
    Splits {tracking_number: [entries]} into (to_check, skipped) based on the cache:
    a tracking number is skipped only if the cache already has it marked "Delivered".
    """
    to_check = {}
    skipped = {}
    for tracking, entries in grouped.items():
        cached = cache.get(tracking, {})
        if cached.get("status") == "Delivered":
            skipped[tracking] = entries
        else:
            to_check[tracking] = entries
    return to_check, skipped


def format_check_tracking_summary(result: CheckTrackingResult) -> str:
    """Plain-text console summary, mirroring verify_tracking.format_verify_summary's style."""
    lines = [
        "=" * 60,
        "CHECK TRACKING SUMMARY",
        "=" * 60,
        f"Checked:           {result.checked}",
        f"Skipped (delivered): {result.skipped_delivered}",
        f"Errors:            {result.errors}",
        f"Updated at:        {result.updated_at}",
        "=" * 60,
    ]
    return "\n".join(lines)


def run_check_tracking(config: dict) -> CheckTrackingResult:
    """
    Standalone entry point for --check-tracking. Visits carrier tracking pages for
    every shipment's tracking number (one visit per unique tracking number, skipping
    any already cached as "Delivered"), and persists results to the tracking-status
    cache workbook after every check so progress survives a mid-run crash.
    """
    from upload_tracking import create_browser_context

    cache_path = config.get("tracking_status_cache", STATUS_CACHE_PATH_DEFAULT)

    entries = build_check_list(config)
    grouped = group_by_tracking(entries)
    cache = load_status_cache(cache_path)

    # Trust the supplier's own "delivered on ..." notes (column K) over a carrier
    # scrape where available — skips the carrier entirely for those.
    cache = seed_delivered_from_notes(grouped, cache)
    save_status_cache(cache_path, cache)

    to_check, skipped = partition_by_delivered(grouped, cache)

    logger.info(f"Check tracking: {len(to_check)} to check, {len(skipped)} skipped (already delivered)")

    checked = 0
    errors = 0

    if to_check:
        playwright, context = create_browser_context(config)
        try:
            page = context.new_page()
            for tracking, group_entries in to_check.items():
                primary = group_entries[0]
                logger.info(f"  Checking {tracking} ({primary.get('carrier')}) — {len(group_entries)} shipment row(s)")
                result = check_tracking_status(
                    page, tracking, primary.get("carrier", ""), config.get("logs_folder")
                )
                if result["status"] == "Unknown":
                    errors += 1

                now = datetime.now().strftime("%Y-%m-%d %H:%M")
                cache[tracking] = {
                    "carrier": primary.get("carrier", ""),
                    "fba_id": primary.get("fba_id", ""),
                    "name": primary.get("name", ""),
                    "destination": primary.get("destination", ""),
                    "ctns": primary.get("ctns", ""),
                    "shipping_way": primary.get("shipping_way", ""),
                    "notes": primary.get("notes", ""),
                    "label_created_date": result["label_created_date"],
                    "expected_delivery_date": result["expected_delivery_date"],
                    "status": result["status"],
                    "last_checked": now,
                    "region": primary.get("region", ""),
                }
                checked += 1
                save_status_cache(cache_path, cache)
        finally:
            try:
                context.close()
                playwright.stop()
            except Exception:
                pass

    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    return CheckTrackingResult(
        checked=checked, skipped_delivered=len(skipped), errors=errors, updated_at=updated_at
    )
