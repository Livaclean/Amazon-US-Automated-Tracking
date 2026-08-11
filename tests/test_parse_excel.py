import pytest
import os
import sys
from pathlib import Path
import xlrd
import logging
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from parse_excel import (
    detect_excel_engine,
    load_us_fc_prefixes,
    is_us_fc,
    group_by_fba_id,
    categorize_shipments,
    load_fc_prefixes,
    is_region_fc,
    parse_and_filter_by_region,
    parse_and_filter_by_region_full,
    find_excel_files,
    load_excel_file,
    parse_and_filter,
    _detect_xls_sheet_cols,
)


def test_categorize_splits_missing_from_has_tracking():
    grouped = {
        "FBA001": [{"tracking": "1Z123", "carrier": "UPS", "row_number": 2}],
        "FBA002": [{"tracking": "", "carrier": "", "row_number": 3}],
        "FBA003": [
            {"tracking": "1ZABC", "carrier": "UPS", "row_number": 4},
            {"tracking": "", "carrier": "", "row_number": 5},
        ],
    }
    has, missing = categorize_shipments(grouped)
    assert "FBA001" in has
    assert "FBA002" in missing
    assert "FBA003" in has   # has at least one tracking entry

def test_slash_tracking_excluded_from_has_tracking():
    grouped = {
        "FBA004": [{"tracking": "1Z123/456", "carrier": "UPS", "row_number": 2}],
    }
    has, missing = categorize_shipments(grouped)
    assert "FBA004" in missing  # only tracking had "/" — treated as missing

def test_load_excel_file_includes_rows_with_empty_tracking(tmp_path):
    """Rows with FBA ID but no tracking should still appear."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B", "C", "D_fc", "E_fba", "F", "G", "H_tracking", "I_carrier"])
    ws.append([None, None, None, "BNA1", "FBA999", None, None, "", ""])
    path = tmp_path / "test.xlsx"
    wb.save(path)
    config = {
        "column_fc_code": 3, "column_fba_id": 4,
        "column_tracking": 7, "column_carrier": 8,
    }
    from parse_excel import load_excel_file
    rows = load_excel_file(str(path), config)
    assert any(r["fba_id"] == "FBA999" for r in rows)


def test_detect_engine_xls():
    assert detect_excel_engine("shipments.xls") == "xlrd"

def test_detect_engine_xlsx():
    assert detect_excel_engine("shipments.xlsx") == "openpyxl"

def test_detect_engine_case_insensitive():
    assert detect_excel_engine("SHIPMENTS.XLS") == "xlrd"
    assert detect_excel_engine("SHIPMENTS.XLSX") == "openpyxl"


def test_load_us_fc_prefixes(tmp_path):
    fc_file = tmp_path / "us_fc_codes.txt"
    fc_file.write_text("# comment\nBNA\nPHX\n\nRIC\n")
    result = load_us_fc_prefixes(str(fc_file))
    assert result == {"BNA", "PHX", "RIC"}

def test_load_us_fc_prefixes_empty(tmp_path):
    fc_file = tmp_path / "us_fc_codes.txt"
    fc_file.write_text("# only comments\n\n")
    assert load_us_fc_prefixes(str(fc_file)) == set()


def test_is_us_fc_match():
    assert is_us_fc("BNA6", {"BNA", "PHX"}) is True

def test_is_us_fc_no_match():
    assert is_us_fc("YYZ1", {"BNA", "PHX"}) is False

def test_is_us_fc_case_insensitive():
    assert is_us_fc("bna6", {"BNA"}) is True

def test_is_us_fc_too_short():
    assert is_us_fc("BN", {"BNA"}) is False

def test_is_us_fc_none():
    assert is_us_fc(None, {"BNA"}) is False

def test_is_us_fc_strips_whitespace():
    assert is_us_fc("  BNA6  ", {"BNA"}) is True


def test_group_by_fba_id_basic():
    rows = [
        {"fba_id": "FBA123", "tracking_num": "1Z001", "carrier": "UPS"},
        {"fba_id": "FBA123", "tracking_num": "1Z002", "carrier": "UPS"},
        {"fba_id": "FBA456", "tracking_num": "999001", "carrier": "FedEx"},
    ]
    result = group_by_fba_id(rows)
    assert result["FBA123"] == [
        {"tracking": "1Z001", "carrier": "UPS", "row_number": None},
        {"tracking": "1Z002", "carrier": "UPS", "row_number": None},
    ]
    assert result["FBA456"] == [{"tracking": "999001", "carrier": "FedEx", "row_number": None}]

def test_group_by_fba_id_deduplication():
    rows = [
        {"fba_id": "FBA123", "tracking_num": "1Z001", "carrier": "UPS"},
        {"fba_id": "FBA123", "tracking_num": "1Z001", "carrier": "UPS"},
    ]
    result = group_by_fba_id(rows)
    assert len(result["FBA123"]) == 1

def test_group_by_fba_id_skips_empty_fba():
    rows = [
        {"fba_id": "", "tracking_num": "1Z001", "carrier": "UPS"},
        {"fba_id": None, "tracking_num": "1Z002", "carrier": "UPS"},
        {"fba_id": "FBA123", "tracking_num": "1Z003", "carrier": "UPS"},
    ]
    result = group_by_fba_id(rows)
    assert list(result.keys()) == ["FBA123"]


def test_load_fc_prefixes_returns_set(tmp_path):
    f = tmp_path / "codes.txt"
    f.write_text("BNA\nPHX\n# comment\n\nIND\n")
    result = load_fc_prefixes(str(f))
    assert result == {"BNA", "PHX", "IND"}


def test_load_fc_prefixes_missing_file(tmp_path):
    result = load_fc_prefixes(str(tmp_path / "missing.txt"))
    assert result == set()


def test_load_fc_prefixes_handles_4letter_codes(tmp_path):
    # PRTO is a 4-letter Canadian prefix — should be stored as-is
    f = tmp_path / "ca.txt"
    f.write_text("YVR\nYYZ\nPRTO\n")
    result = load_fc_prefixes(str(f))
    assert "PRTO" in result


def test_parse_and_filter_by_region_returns_dict_keyed_by_region(tmp_path):
    # Build minimal config pointing to tmp fc_codes files
    us_file = tmp_path / "us.txt"
    ca_file = tmp_path / "ca.txt"
    us_file.write_text("BNA\n")
    ca_file.write_text("YVR\n")

    config = {
        "input_folder": str(tmp_path / "input"),
        "column_fc_code": 0,
        "column_fba_id": 1,
        "column_tracking": 2,
        "column_carrier": 3,
        "regions": [
            {"name": "US", "amazon_url": "https://sellercentral.amazon.com", "fc_codes_file": str(us_file)},
            {"name": "CA", "amazon_url": "https://sellercentral.amazon.ca",  "fc_codes_file": str(ca_file)},
        ],
    }

    # No Excel file — should return empty dicts for each region
    Path(config["input_folder"]).mkdir()
    result = parse_and_filter_by_region(config)
    assert "US" in result
    assert "CA" in result
    assert result["US"] == {}
    assert result["CA"] == {}


# ---------------------------------------------------------------------------
# NEW: find_excel_files, load_excel_file, parse_and_filter pipeline
# ---------------------------------------------------------------------------

def test_find_excel_files_finds_both_types(tmp_path):
    (tmp_path / "a.xlsx").write_text("")
    (tmp_path / "b.xls").write_text("")
    result = find_excel_files(str(tmp_path))
    assert len(result) == 2
    names = [Path(f).name for f in result]
    assert "a.xlsx" in names
    assert "b.xls" in names


def test_find_excel_files_empty_folder(tmp_path):
    assert find_excel_files(str(tmp_path)) == []


def test_find_excel_files_ignores_csv(tmp_path):
    (tmp_path / "data.csv").write_text("")
    (tmp_path / "data.xlsx").write_text("")
    result = find_excel_files(str(tmp_path))
    assert len(result) == 1
    assert result[0].endswith(".xlsx")


def test_load_excel_file_xls_real(sample_xls):
    config = {"column_fc_code": 3, "column_fba_id": 4,
              "column_tracking": 7, "column_carrier": 8}
    rows = load_excel_file(sample_xls, config)
    assert len(rows) > 0
    assert all("fba_id" in r for r in rows)


def test_load_excel_file_xls_numeric_cells(sample_xls):
    config = {"column_fc_code": 3, "column_fba_id": 4,
              "column_tracking": 7, "column_carrier": 8}
    rows = load_excel_file(sample_xls, config)
    for r in rows:
        if r["tracking_num"]:
            assert ".0" not in r["tracking_num"], f"Found '.0' in tracking: {r['tracking_num']}"


def test_load_excel_file_xlsx_multi_sheet(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["A", "B", "C", "D_fc", "E_fba", "F", "G", "H_tracking", "I_carrier"])
    ws1.append([None, None, None, "BNA6", "FBA_S1", None, None, "1Z001", "UPS"])
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["A", "B", "C", "D_fc", "E_fba", "F", "G", "H_tracking", "I_carrier"])
    ws2.append([None, None, None, "GYR3", "FBA_S2", None, None, "1Z002", "UPS"])
    path = tmp_path / "multi.xlsx"
    wb.save(path)
    config = {"column_fc_code": 3, "column_fba_id": 4,
              "column_tracking": 7, "column_carrier": 8}
    rows = load_excel_file(str(path), config)
    fba_ids = {r["fba_id"] for r in rows}
    assert "FBA_S1" in fba_ids
    assert "FBA_S2" in fba_ids


def test_parse_and_filter_full_pipeline(tmp_config):
    import openpyxl as xl
    wb = xl.Workbook()
    ws = wb.active
    ws.append(["A", "B", "C", "D_fc", "E_fba", "F", "G", "H_tracking", "I_carrier"])
    ws.append([None, None, None, "BNA6", "FBA_P1", None, None, "1ZTEST001", "UPS"])
    ws.append([None, None, None, "YVR2", "FBA_P2", None, None, "1ZTEST002", "UPS"])
    path = Path(tmp_config["input_folder"]) / "test.xlsx"
    wb.save(path)
    result = parse_and_filter(tmp_config)
    assert "FBA_P1" in result
    assert "FBA_P2" not in result


def test_parse_and_filter_no_files(tmp_config):
    result = parse_and_filter(tmp_config)
    assert result == {}


def test_parse_and_filter_by_region_with_data(tmp_config):
    import openpyxl as xl
    wb = xl.Workbook()
    ws = wb.active
    ws.append(["A", "B", "C", "D_fc", "E_fba", "F", "G", "H_tracking", "I_carrier"])
    ws.append([None, None, None, "BNA6", "FBA_US1", None, None, "1Z001", "UPS"])
    ws.append([None, None, None, "YVR2", "FBA_CA1", None, None, "1Z002", "UPS"])
    path = Path(tmp_config["input_folder"]) / "regions.xlsx"
    wb.save(path)
    result = parse_and_filter_by_region(tmp_config)
    assert "FBA_US1" in result.get("US", {})
    assert "FBA_CA1" in result.get("CA", {})
    assert "FBA_US1" not in result.get("CA", {})


def test_group_by_fba_id_slash_split():
    rows = [{"fba_id": "STAR-A/STAR-B", "tracking_num": "1Z001", "carrier": "UPS", "row_number": 2}]
    result = group_by_fba_id(rows)
    assert "STAR-A" in result
    assert "STAR-B" in result
    assert result["STAR-A"][0]["tracking"] == "1Z001"


def test_group_by_fba_id_skips_walmart_and_tiktok_ids():
    rows = [
        {"fba_id": "WMT123WFA", "tracking_num": "1Z001", "carrier": "UPS"},
        {"fba_id": "IBR5766196157831222022", "tracking_num": "1Z002", "carrier": "UPS"},
        {"fba_id": "STAR-A/IBR999", "tracking_num": "1Z003", "carrier": "UPS"},
        {"fba_id": "FBA123", "tracking_num": "1Z004", "carrier": "UPS"},
    ]
    result = group_by_fba_id(rows)
    assert list(result.keys()) == ["STAR-A", "FBA123"]


def test_parse_and_filter_by_region_full_returns_unmatched_rows(tmp_path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B", "C", "D_fc", "E_fba", "F", "G", "H_tracking", "I_carrier"])
    ws.append([None, None, None, "BNA6", "FBA_US", None, None, "1Z001", "UPS"])
    ws.append([None, None, None, "ZZZ9", "FBA_UNKNOWN", None, None, "1Z002", "UPS"])

    input_dir = tmp_path / "input"
    input_dir.mkdir()
    wb.save(input_dir / "test.xlsx")

    us_fc = tmp_path / "us.txt"
    us_fc.write_text("BNA\n")
    config = {
        "input_folder": str(input_dir),
        "column_fc_code": 3, "column_fba_id": 4,
        "column_tracking": 7, "column_carrier": 8,
        "regions": [{"name": "US", "amazon_url": "https://x", "fc_codes_file": str(us_fc)}],
    }

    from parse_excel import parse_and_filter_by_region_full
    region_dict, unmatched = parse_and_filter_by_region_full(config)

    assert "FBA_US" in region_dict["US"]
    assert len(unmatched) == 1
    assert unmatched[0]["fba_id"] == "FBA_UNKNOWN"
    assert unmatched[0]["fc_code"] == "ZZZ9"


def test_parse_and_filter_by_region_still_returns_region_dict_only(tmp_path):
    """parse_and_filter_by_region() must keep its existing return shape — no unmatched_rows leak in."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B", "C", "D_fc", "E_fba", "F", "G", "H_tracking", "I_carrier"])
    ws.append([None, None, None, "BNA6", "FBA_US", None, None, "1Z001", "UPS"])

    input_dir = tmp_path / "input"
    input_dir.mkdir()
    wb.save(input_dir / "test.xlsx")

    us_fc = tmp_path / "us.txt"
    us_fc.write_text("BNA\n")
    config = {
        "input_folder": str(input_dir),
        "column_fc_code": 3, "column_fba_id": 4,
        "column_tracking": 7, "column_carrier": 8,
        "regions": [{"name": "US", "amazon_url": "https://x", "fc_codes_file": str(us_fc)}],
    }

    from parse_excel import parse_and_filter_by_region
    region_dict = parse_and_filter_by_region(config)
    assert isinstance(region_dict, dict)
    assert "FBA_US" in region_dict["US"]


# ---------------------------------------------------------------------------
# NEW: _detect_xls_sheet_cols extended column detection tests
# ---------------------------------------------------------------------------

class _FakeXlrdCell:
    def __init__(self, value):
        self.value = value
        self.ctype = xlrd.XL_CELL_NUMBER if isinstance(value, (int, float)) else xlrd.XL_CELL_TEXT


class _FakeXlrdSheet:
    def __init__(self, name, rows):
        self.name = name
        self._rows = rows
        self.nrows = len(rows)
        self.ncols = max((len(r) for r in rows), default=0)

    def cell(self, r, c):
        row = self._rows[r]
        return _FakeXlrdCell(row[c] if c < len(row) else "")


def test_detect_xls_sheet_cols_us_shape():
    """11 columns, separate ITEMS column, blank last-column header — matches the real US sheet."""
    header = ["SYSTEM NO", "Order No", "ITEMS", "DESTINATION", "FBA ID",
              "NO OF CTNS ", "SHIPPING  WAY", "TRACKING NUMBERS", "CARRIER", "ETD", ""]
    data = ["A251014HX059", "Widget Variety Pack", "", "ORF2", "FBA1924FWPYT",
            9, "express", "1ZA6D7510412465060", "UPS", "", "delivered on 2026.02.24"]
    sheet = _FakeXlrdSheet("US", [header, data])
    cols = _detect_xls_sheet_cols(sheet)
    assert cols == {
        "header_row": 0, "col_fc": 3, "col_fba": 4, "col_tracking": 7,
        "col_carrier": 8, "col_name": 1, "col_ctns": 5, "col_shipping_way": 6,
        "col_notes": 10,
    }


def test_detect_xls_sheet_cols_de_shape():
    """10 columns, merged 'Order No-ITEMS' column, named 'ETAs' last column — matches the real DE sheet."""
    header = ["SYSTEM NO", "Order No-ITEMS", "DESTINATION", "FBA ID",
              "NO OF CTNS ", "SHIPPING  WAY", "TRACKING NUMBERS", "CARRIER", "ETD", "ETAs"]
    data = ["A250710HX090", "Widget DE", "DMT2", "FBA15KK5TKDF",
            4, "C-AIR", "1ZC51W066825252132", "ups", "", "delivered on 7.31"]
    sheet = _FakeXlrdSheet("DE", [header, data])
    cols = _detect_xls_sheet_cols(sheet)
    assert cols == {
        "header_row": 0, "col_fc": 2, "col_fba": 3, "col_tracking": 6,
        "col_carrier": 7, "col_name": 1, "col_ctns": 4, "col_shipping_way": 5,
        "col_notes": 9,
    }


def test_detect_xls_sheet_cols_falls_back_with_warning_when_field_missing(caplog):
    """Header row has FBA ID + TRACKING but no recognizable name/ctns/shipping_way labels."""
    header = ["SYSTEM NO", "X", "Y", "FBA ID", "Z", "TRACKING NUMBERS", "CARRIER"]
    data = ["A1", "b", "c", "FBA001", "d", "1Z001", "UPS"]
    sheet = _FakeXlrdSheet("ODD", [header, data])
    with caplog.at_level(logging.WARNING):
        cols = _detect_xls_sheet_cols(sheet)
    assert cols["col_name"] == 1
    assert cols["col_ctns"] == 5
    assert cols["col_shipping_way"] == 6
    assert cols["col_notes"] == 6  # last column (ncols=7, so index 6)
    assert "ODD" in caplog.text
    assert "name" in caplog.text.lower()
    assert "ctns" in caplog.text.lower()
    assert "shipping_way" in caplog.text.lower()


def test_detect_xls_sheet_cols_full_fallback_when_no_header_found():
    """No row in the first 3 has both 'FBA ID' and 'TRACKING' — full default fallback, unchanged behavior."""
    sheet = _FakeXlrdSheet("WEIRD", [["nothing", "here", "matches"]])
    cols = _detect_xls_sheet_cols(sheet)
    assert cols["header_row"] == 0
    assert cols["col_fc"] == 3
    assert cols["col_fba"] == 4
    assert cols["col_tracking"] == 7
    assert cols["col_carrier"] == 8
    assert cols["col_name"] == 1
    assert cols["col_ctns"] == 5
    assert cols["col_shipping_way"] == 6
    assert cols["col_notes"] == 2  # last column (ncols=3, so index 2)
