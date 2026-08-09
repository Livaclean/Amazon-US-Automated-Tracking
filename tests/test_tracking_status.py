import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import pytest
import openpyxl
import xlrd

from tracking_status import (
    load_row_context,
    build_check_list,
    group_by_tracking,
    load_status_cache,
    save_status_cache,
    partition_by_delivered,
    format_check_tracking_summary,
    CheckTrackingResult,
    STATUS_CACHE_COLUMNS,
    _extract_status_from_text,
    _extract_status_and_dates_from_text,
    _notes_indicate_delivered,
    _extract_date_from_notes,
    seed_delivered_from_notes,
    _fedex_status_from_api,
    _dhl_status_from_api,
    _row_context_from_xls_book,
)

# Real UPS tracking-detail page text (captured live) for a shipment whose label
# was created but not yet picked up by UPS. "Out for Delivery" and "Delivery" are
# listed as *future* inactive milestones on every UPS page regardless of the real
# current status — only "Label Created" is marked "active" here.
REAL_UPS_LABEL_CREATED_PAGE = """
Tracking Details
Tracking No. or Delivery Notice
Label Created
1Z035B4X0338736309

Latest Update

Estimated delivery date will be available when UPS receives the package.

Label Created
active
We Have Your Package
inactive
On the Way
inactive
Out for Delivery
inactive
Delivery
inactive
Show Details

Shipment Details
1 of 2 Piece Shipment

Ship To

LENEXA, KS US

Service

UPS Ground

Shipped / Billed On

07/16/2026
"""

# Real UPS tracking-detail page text (captured live) for an already-delivered
# shipment. UPS shows no year — just "Friday, July 17" — and no "Delivered On"
# label, just the bare date/time sitting right after the "Delivered" banner.
REAL_UPS_DELIVERED_PAGE = """
Tracking Details
Tracking No. or Delivery Notice
Delivered check_circle
1Z035B4X0315994412 content_copy
Copy Tracking Number
check
close

Tracking number copied to clipboard.

Friday, July 17
at
10:20 A.M.
Delivered To
RIALTO, CA US
Received By
SOF AMZ LIN
Show Details

Shipment Details
1 of 3 Piece Shipment

Delivered To

RIALTO, CA US

Service

UPS Ground

Shipped / Billed On

06/20/2026
"""

# Real FedEx tracking-detail page text (captured live) for a shipment whose
# label was created but not yet picked up. Unlike UPS, the milestone progress
# bar ("WE HAVE YOUR PACKAGE" / "ON THE WAY" / "OUT FOR DELIVERY") has no
# active/inactive markers at all — every step is always shown as plain text.
# The real current status only appears once, directly under "FROM".
REAL_FEDEX_LABEL_CREATED_PAGE = """
Fedex Home
Shipping
Tracking
Sign Up or Log In
Tracking ID:
874466268456
Local Scan Time
DELIVERY DETAILS

We'll add a delivery date as soon as we get your package.
Shipment is 1 of 3 pieces

Get updates
Email address *
SUBMIT
MORE OPTIONS

FROM
Label created
Rancho Cucamonga, CA US
View more details

WE HAVE YOUR PACKAGE

ON THE WAY

OUT FOR DELIVERY

TO
HUNTLEY, IL US
"""


CONTEXT_CONFIG = {
    "column_fc_code": 3,
    "column_name": 1,
    "column_ctns": 5,
    "column_shipping_way": 6,
    "column_notes": 10,
}


def _write_context_sheet(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["SYSTEM NO", "Order No", "ITEMS", "DESTINATION", "FBA ID",
               "NO OF CTNS", "SHIPPING WAY", "TRACKING NUMBERS", "CARRIER", "ETD", None])
    ws.append(["A1", "Widget Variety Pack", None, "ORF2", "FBA001",
               9, "express", "1Z001", "UPS", None, "delivered on 2026.02.24"])
    ws.append(["A2", "Pimple Patches", None, "IUSF", "FBA002",
               4, "C-SEA", "1Z002", "UPS", None, None])
    path = tmp_path / "shipments.xlsx"
    wb.save(path)
    return str(path)


@pytest.mark.unit
def test_load_row_context_extracts_descriptive_columns(tmp_path):
    path = _write_context_sheet(tmp_path)
    ctx = load_row_context(path, CONTEXT_CONFIG)
    assert ctx["FBA001"]["name"] == "Widget Variety Pack"
    assert ctx["FBA001"]["destination"] == "ORF2"
    assert ctx["FBA001"]["ctns"] == 9
    assert ctx["FBA001"]["shipping_way"] == "express"
    assert ctx["FBA001"]["notes"] == "delivered on 2026.02.24"


@pytest.mark.unit
def test_load_row_context_handles_blank_notes(tmp_path):
    path = _write_context_sheet(tmp_path)
    ctx = load_row_context(path, CONTEXT_CONFIG)
    assert ctx["FBA002"]["notes"] == ""


class _FakeXlrdCell:
    def __init__(self, value):
        self.value = value
        self.ctype = xlrd.XL_CELL_NUMBER if isinstance(value, (int, float)) else xlrd.XL_CELL_TEXT


class _FakeXlrdSheet:
    def __init__(self, name, rows):
        self.name = name
        self._rows = rows
        self.nrows = len(rows)
        self.ncols = max((len(r) for r in rows), default=0)

    def cell(self, r, c):
        row = self._rows[r]
        return _FakeXlrdCell(row[c] if c < len(row) else "")


class _FakeXlrdBook:
    def __init__(self, sheets):
        self._sheets = sheets
        self.nsheets = len(sheets)

    def sheet_by_index(self, i):
        return self._sheets[i]


@pytest.mark.unit
def test_row_context_from_xls_book_us_shape():
    header = ["SYSTEM NO", "Order No", "ITEMS", "DESTINATION", "FBA ID",
              "NO OF CTNS ", "SHIPPING  WAY", "TRACKING NUMBERS", "CARRIER", "ETD", ""]
    data = ["A251014HX059", "Widget Variety Pack", "", "ORF2", "FBA1924FWPYT",
            9, "express", "1ZA6D7510412465060", "UPS", "", "delivered on 2026.02.24"]
    wb = _FakeXlrdBook([_FakeXlrdSheet("US", [header, data])])
    ctx = _row_context_from_xls_book(wb)
    assert ctx["FBA1924FWPYT"] == {
        "name": "Widget Variety Pack",
        "destination": "ORF2",
        "ctns": "9",
        "shipping_way": "express",
        "notes": "delivered on 2026.02.24",
    }


@pytest.mark.unit
def test_row_context_from_xls_book_de_shape():
    """10-column sheet, merged Order No-ITEMS, named ETAs last column — column positions differ from US."""
    header = ["SYSTEM NO", "Order No-ITEMS", "DESTINATION", "FBA ID",
              "NO OF CTNS ", "SHIPPING  WAY", "TRACKING NUMBERS", "CARRIER", "ETD", "ETAs"]
    data = ["A250710HX090", "Widget DE", "DMT2", "FBA15KK5TKDF",
            4, "C-AIR", "1ZC51W066825252132", "ups", "", "delivered on 7.31"]
    wb = _FakeXlrdBook([_FakeXlrdSheet("DE", [header, data])])
    ctx = _row_context_from_xls_book(wb)
    assert ctx["FBA15KK5TKDF"] == {
        "name": "Widget DE",
        "destination": "DMT2",
        "ctns": "4",
        "shipping_way": "C-AIR",
        "notes": "delivered on 7.31",
    }


@pytest.mark.unit
def test_row_context_from_xls_book_cross_sheet_row_number_collision_resolved_by_fba_id():
    """Two sheets both have a physical 'row 2' with different FBA IDs — proves the join no longer
    collides now that it's keyed by FBA ID instead of the (non-unique-across-sheets) row_number."""
    header = ["SYSTEM NO", "Order No", "ITEMS", "DESTINATION", "FBA ID",
              "NO OF CTNS ", "SHIPPING  WAY", "TRACKING NUMBERS", "CARRIER", "ETD", ""]
    row_a = ["A1", "Sheet A Item", "", "ORF2", "FBA001", 1, "air", "1Z001", "UPS", "", ""]
    row_b = ["A2", "Sheet B Item", "", "ORF3", "FBA002", 2, "sea", "1Z002", "UPS", "", ""]
    wb = _FakeXlrdBook([
        _FakeXlrdSheet("SheetA", [header, row_a]),
        _FakeXlrdSheet("SheetB", [header, row_b]),
    ])
    ctx = _row_context_from_xls_book(wb)
    assert ctx["FBA001"]["name"] == "Sheet A Item"
    assert ctx["FBA002"]["name"] == "Sheet B Item"


@pytest.mark.unit
def test_load_row_context_dispatches_to_xls_reader_by_extension(monkeypatch, tmp_path):
    """load_row_context() picks the xlrd path for a .xls filename without opening a real file."""
    called = {}

    def fake_open_workbook(path):
        called["path"] = path
        header = ["SYSTEM NO", "Order No", "ITEMS", "DESTINATION", "FBA ID",
                  "NO OF CTNS ", "SHIPPING  WAY", "TRACKING NUMBERS", "CARRIER", "ETD", ""]
        data = ["A1", "Widget", "", "ORF2", "FBA001", 1, "air", "1Z001", "UPS", "", ""]
        return _FakeXlrdBook([_FakeXlrdSheet("US", [header, data])])

    monkeypatch.setattr(xlrd, "open_workbook", fake_open_workbook)
    fake_path = str(tmp_path / "shipments.xls")
    ctx = load_row_context(fake_path, {})
    assert called["path"] == fake_path
    assert ctx["FBA001"]["name"] == "Widget"


@pytest.mark.unit
def test_group_by_tracking_groups_multiple_fba_rows_under_one_number():
    entries = [
        {"fba_id": "FBA001", "tracking": "1Z001", "carrier": "UPS"},
        {"fba_id": "FBA002", "tracking": "1Z001", "carrier": "UPS"},
        {"fba_id": "FBA003", "tracking": "1Z002", "carrier": "UPS"},
    ]
    grouped = group_by_tracking(entries)
    assert set(grouped.keys()) == {"1Z001", "1Z002"}
    assert len(grouped["1Z001"]) == 2
    assert len(grouped["1Z002"]) == 1


@pytest.mark.unit
def test_group_by_tracking_skips_blank_tracking():
    entries = [
        {"fba_id": "FBA001", "tracking": "", "carrier": "UPS"},
        {"fba_id": "FBA002", "tracking": "1Z002", "carrier": "UPS"},
    ]
    grouped = group_by_tracking(entries)
    assert list(grouped.keys()) == ["1Z002"]


@pytest.mark.unit
def test_build_check_list_merges_shipments_with_row_context(tmp_config, tmp_path):
    import openpyxl as xl
    wb = xl.Workbook()
    ws = wb.active
    ws.append(["SYSTEM NO", "Order No", "ITEMS", "DESTINATION", "FBA ID",
               "NO OF CTNS", "SHIPPING WAY", "TRACKING NUMBERS", "CARRIER", "ETD", None])
    ws.append(["A1", "Widget Variety Pack", None, "BNA6", "FBA_CL1",
               9, "express", "1ZCL001", "UPS", None, None])
    from pathlib import Path
    path = Path(tmp_config["input_folder"]) / "check_list.xlsx"
    wb.save(path)
    tmp_config.update(CONTEXT_CONFIG)

    entries = build_check_list(tmp_config)
    assert len(entries) == 1
    e = entries[0]
    assert e["fba_id"] == "FBA_CL1"
    assert e["tracking"] == "1ZCL001"
    assert e["carrier"] == "UPS"
    assert e["region"] == "US"
    assert e["name"] == "Widget Variety Pack"
    assert e["destination"] == "BNA6"
    assert e["ctns"] == 9
    assert e["shipping_way"] == "express"


@pytest.mark.unit
def test_build_check_list_skips_blank_tracking(tmp_config):
    import openpyxl as xl
    from pathlib import Path
    wb = xl.Workbook()
    ws = wb.active
    ws.append(["SYSTEM NO", "Order No", "ITEMS", "DESTINATION", "FBA ID",
               "NO OF CTNS", "SHIPPING WAY", "TRACKING NUMBERS", "CARRIER", "ETD", None])
    ws.append(["A1", "No Tracking Yet", None, "BNA6", "FBA_CL2",
               1, "sea", "", "", None, None])
    path = Path(tmp_config["input_folder"]) / "check_list2.xlsx"
    wb.save(path)
    tmp_config.update(CONTEXT_CONFIG)

    entries = build_check_list(tmp_config)
    assert entries == []


@pytest.mark.unit
def test_status_cache_round_trip(tmp_path):
    path = str(tmp_path / "tracking_status.xlsx")
    cache = {
        "1Z001": {
            "carrier": "UPS", "fba_id": "FBA001", "name": "Widget Pack",
            "destination": "ORF2", "ctns": 9, "shipping_way": "express",
            "notes": "", "label_created_date": "2026-06-01",
            "expected_delivery_date": "2026-06-10", "status": "In Transit",
            "last_checked": "2026-06-05 10:00", "region": "US",
        }
    }
    save_status_cache(path, cache)
    loaded = load_status_cache(path)
    assert loaded["1Z001"]["status"] == "In Transit"
    assert loaded["1Z001"]["fba_id"] == "FBA001"
    assert loaded["1Z001"]["ctns"] == 9


@pytest.mark.unit
def test_load_status_cache_missing_file_returns_empty_dict(tmp_path):
    path = str(tmp_path / "does_not_exist.xlsx")
    assert load_status_cache(path) == {}


@pytest.mark.unit
def test_status_cache_columns_header_order():
    assert STATUS_CACHE_COLUMNS[0] == "Tracking Number"
    assert "Current Status" in STATUS_CACHE_COLUMNS
    assert "Last Checked" in STATUS_CACHE_COLUMNS


@pytest.mark.unit
def test_partition_by_delivered_skips_cached_delivered():
    grouped = {"1Z001": [{"fba_id": "FBA001"}], "1Z002": [{"fba_id": "FBA002"}]}
    cache = {"1Z001": {"status": "Delivered"}}
    to_check, skipped = partition_by_delivered(grouped, cache)
    assert list(to_check.keys()) == ["1Z002"]
    assert list(skipped.keys()) == ["1Z001"]


@pytest.mark.unit
def test_partition_by_delivered_rechecks_uncached_and_non_delivered():
    grouped = {"1Z001": [{"fba_id": "FBA001"}], "1Z002": [{"fba_id": "FBA002"}]}
    cache = {"1Z001": {"status": "In Transit"}}
    to_check, skipped = partition_by_delivered(grouped, cache)
    assert set(to_check.keys()) == {"1Z001", "1Z002"}
    assert skipped == {}


@pytest.mark.unit
def test_extract_status_from_text_recognizes_invalid_tracking_number():
    text = "warning\nInvalid tracking number\nThis tracking number may be invalid or not active yet."
    assert _extract_status_from_text(text) == "Invalid"


@pytest.mark.unit
def test_extract_status_uses_active_milestone_not_first_keyword_match():
    """
    Regression test for a live bug: a page listing 'Out for Delivery' as a future
    *inactive* milestone was misreported as status "Out for Delivery" even though
    "Label Created" was the actual active step. The active/inactive markers must
    be used, not a plain "does this phrase appear anywhere on the page" search.
    """
    result = _extract_status_and_dates_from_text(REAL_UPS_LABEL_CREATED_PAGE)
    assert result["status"] == "Label Created"


@pytest.mark.unit
def test_extract_label_created_date_from_shipped_billed_on():
    result = _extract_status_and_dates_from_text(REAL_UPS_LABEL_CREATED_PAGE)
    assert result["label_created_date"] == "07/16/2026"


@pytest.mark.unit
def test_extract_status_and_actual_date_for_delivered_ups_page():
    """
    Delivered UPS pages don't use the active/inactive milestone list or a
    'Scheduled Delivery' label at all — just a bare weekday/month/day near the
    'Delivered' banner (no year). Both status and that date must be captured.
    """
    result = _extract_status_and_dates_from_text(REAL_UPS_DELIVERED_PAGE)
    assert result["status"] == "Delivered"
    assert result["expected_delivery_date"] == "Friday, July 17"


@pytest.mark.unit
def test_extract_status_for_fedex_uses_from_line_not_milestone_bar():
    """
    Regression test for a live bug: FedEx always lists every milestone
    ("OUT FOR DELIVERY" included) as plain text with no active/inactive markers,
    so a whole-page keyword scan wrongly reports "Out for Delivery" even when the
    real status ("Label created") sits right under "FROM".
    """
    result = _extract_status_and_dates_from_text(REAL_FEDEX_LABEL_CREATED_PAGE)
    assert result["status"] == "Label Created"


# Trimmed from a real api.fedex.com/track/v2/shipments response (captured live
# via network interception) for a label-created-but-not-picked-up shipment.
REAL_FEDEX_API_LABEL_CREATED = {
    "output": {
        "packages": [{
            "keyStatus": "Label created",
            "lastScanStatus": "Shipment information sent to FedEx",
            "lastScanDateTime": "2026-07-16T03:33:00-07:00",
            "scanEventList": [{
                "date": "2026-07-16", "time": "03:33:00", "gmtOffset": "-07:00",
                "status": "Shipment information sent to FedEx", "statusCD": "OC",
            }],
            "displayShipDt": "Will be updated soon",
            "displayEstDeliveryDt": "Will be updated soon",
            "displayActDeliveryDt": "",
        }]
    }
}


@pytest.mark.unit
def test_fedex_status_from_api_extracts_label_created_date():
    result = _fedex_status_from_api(REAL_FEDEX_API_LABEL_CREATED)
    assert result["status"] == "Label Created"
    assert result["label_created_date"] == "2026-07-16"
    assert result["expected_delivery_date"] is None


@pytest.mark.unit
def test_fedex_status_from_api_no_packages_returns_unknown():
    result = _fedex_status_from_api({"output": {"packages": []}})
    assert result["status"] == "Unknown"


# Trimmed from a real DHL utapi response (captured live via network interception)
# for a delivered shipment. Several transit-event descriptions contain the word
# "shipment" (e.g. "Shipment is out with courier for delivery") — a naive keyword
# scan over `events` picks one of those instead of the real label-created date,
# which only `details.shipmentActivationDate` reliably gives.
REAL_DHL_API_DELIVERED = {
    "shipments": [{
        "status": {
            "timestamp": "2026-06-28T13:08:00+10:00",
            "statusCode": "delivered",
            "status": "101",
            "description": "Delivered",
        },
        "details": {
            "shipmentActivationDate": "2026-06-12T15:26:49",
        },
        "events": [
            {"timestamp": "2026-06-28T13:08:00+10:00", "statusCode": "delivered", "description": "Delivered"},
            {"timestamp": "2026-06-28T12:55:00+10:00", "statusCode": "transit",
             "description": "Shipment is out with courier for delivery"},
            {"timestamp": "2026-06-12T15:26:00+08:00", "statusCode": "transit", "description": "Shipment picked up"},
        ],
    }]
}


@pytest.mark.unit
def test_dhl_status_from_api_uses_shipment_activation_date_not_event_keyword_scan():
    """
    Regression test for a live bug: label_created_date came back as
    "2026-06-28T12:55:00+10:00" (a courier-out-for-delivery event, days later)
    because its description contains the word "shipment" and a naive keyword
    scan over events matched it. shipmentActivationDate is the correct source.
    """
    result = _dhl_status_from_api(REAL_DHL_API_DELIVERED)
    assert result["status"] == "Delivered"
    assert result["label_created_date"] == "2026-06-12T15:26:49"


@pytest.mark.unit
def test_dhl_status_from_api_uses_status_timestamp_as_delivery_date_when_delivered():
    result = _dhl_status_from_api(REAL_DHL_API_DELIVERED)
    assert result["expected_delivery_date"] == "2026-06-28T13:08:00+10:00"


@pytest.mark.unit
def test_dhl_status_from_api_no_shipments_returns_unknown():
    result = _dhl_status_from_api({"shipments": []})
    assert result["status"] == "Unknown"


@pytest.mark.unit
def test_notes_indicate_delivered_true():
    assert _notes_indicate_delivered("delivered on 2026.02.24") is True


@pytest.mark.unit
def test_notes_indicate_delivered_false():
    assert _notes_indicate_delivered("waiting for FEDEX updates") is False


@pytest.mark.unit
def test_notes_indicate_delivered_blank():
    assert _notes_indicate_delivered("") is False


@pytest.mark.unit
def test_extract_date_from_notes_dotted_format():
    assert _extract_date_from_notes("delivered on 2026.02.24") == "2026.02.24"


@pytest.mark.unit
def test_extract_date_from_notes_short_format():
    assert _extract_date_from_notes("delivered on 12.12") == "12.12"


@pytest.mark.unit
def test_extract_date_from_notes_no_date():
    assert _extract_date_from_notes("delivered") is None


@pytest.mark.unit
def test_seed_delivered_from_notes_marks_status_and_date():
    grouped = {
        "1Z001": [{
            "fba_id": "FBA001", "carrier": "UPS", "name": "Widget Pack",
            "destination": "ORF2", "ctns": 9, "shipping_way": "express",
            "notes": "delivered on 2026.02.24", "region": "US",
        }]
    }
    cache = seed_delivered_from_notes(grouped, {})
    assert cache["1Z001"]["status"] == "Delivered"
    assert cache["1Z001"]["expected_delivery_date"] == "2026.02.24"
    assert cache["1Z001"]["fba_id"] == "FBA001"


@pytest.mark.unit
def test_seed_delivered_from_notes_ignores_non_delivered_notes():
    grouped = {
        "1Z002": [{
            "fba_id": "FBA002", "carrier": "UPS", "notes": "waiting for FEDEX updates",
        }]
    }
    cache = seed_delivered_from_notes(grouped, {})
    assert cache == {}


@pytest.mark.unit
def test_seed_delivered_from_notes_then_partition_skips_carrier_check():
    grouped = {
        "1Z001": [{"fba_id": "FBA001", "carrier": "UPS", "notes": "delivered on 2026.02.24"}],
        "1Z002": [{"fba_id": "FBA002", "carrier": "UPS", "notes": ""}],
    }
    cache = seed_delivered_from_notes(grouped, {})
    to_check, skipped = partition_by_delivered(grouped, cache)
    assert list(skipped.keys()) == ["1Z001"]
    assert list(to_check.keys()) == ["1Z002"]


@pytest.mark.unit
def test_format_check_tracking_summary_includes_counts():
    result = CheckTrackingResult(checked=3, skipped_delivered=2, errors=1, updated_at="2026-06-05 10:00")
    text = format_check_tracking_summary(result)
    assert "3" in text
    assert "2" in text
    assert "1" in text
    assert "2026-06-05 10:00" in text
