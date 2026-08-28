import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import pytest
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill

def test_highlight_rows_applies_yellow_fill(tmp_path):
    from highlight_excel import highlight_and_save

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B", "C", "D_fc", "E_fba", "F", "G", "H_tracking", "I_carrier"])
    ws.append([None, None, None, "BNA1", "FBA001", None, None, "1Z123", "UPS"])
    ws.append([None, None, None, "BNA1", "FBA002", None, None, "1Z456", "UPS"])
    ws.append([None, None, None, "BNA1", "FBA003", None, None, "1ZABC", "UPS"])
    src = tmp_path / "test.xlsx"
    wb.save(src)

    updated_rows = {2, 4}  # 1-indexed Excel rows
    dest = tmp_path / "output.xlsx"

    highlight_and_save(str(src), str(dest), updated_rows)

    result = openpyxl.load_workbook(dest)
    ws2 = result.active
    yellow = "FFFF00"
    # Check that highlighted rows have yellow fill on at least the first cell
    assert ws2.cell(2, 1).fill.fgColor.rgb[-6:] == yellow  # FBA001 highlighted
    assert ws2.cell(3, 1).fill.fgColor.rgb[-6:] != yellow  # FBA002 not highlighted
    assert ws2.cell(4, 1).fill.fgColor.rgb[-6:] == yellow  # FBA003 highlighted

def test_highlight_saves_to_xlsx_for_xls_dest(tmp_path):
    """When dest has .xls extension, output should be saved as .xlsx instead."""
    from highlight_excel import highlight_and_save
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["header"])
    ws.append(["data"])
    src = tmp_path / "test.xlsx"
    wb.save(src)
    dest = tmp_path / "output.xls"  # request .xls output
    result_path = highlight_and_save(str(src), str(dest), {2})
    assert result_path.endswith(".xlsx")
    assert Path(result_path).exists()

def test_highlight_xls_source(sample_xls, tmp_path):
    """highlight_and_save should handle .xls source, output as .xlsx."""
    from highlight_excel import highlight_and_save
    import openpyxl
    dest = tmp_path / "output.xlsx"
    result_path = highlight_and_save(sample_xls, str(dest), {2, 3})
    assert result_path.endswith(".xlsx")
    assert Path(result_path).exists()
    wb = openpyxl.load_workbook(result_path)
    assert wb.active is not None


def test_highlight_applies_to_matching_row_on_every_sheet(tmp_path):
    """row_num in updated_rows has no sheet identity attached (parse_excel.py
    numbers rows per-sheet, not globally), so highlight_and_save must not
    silently pick only one sheet — it highlights every sheet that has a row
    at that index, and skips sheets too short to have it."""
    from highlight_excel import highlight_and_save
    import openpyxl

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "US"
    ws1.append(["header"])
    ws1.append(["row2-us"])
    ws2 = wb.create_sheet("AU")
    ws2.append(["header"])
    ws2.append(["row2-au"])
    ws2.append(["row3-au"])
    src = tmp_path / "multi.xlsx"
    wb.save(src)
    dest = tmp_path / "out.xlsx"

    highlight_and_save(str(src), str(dest), {2, 3})

    result = openpyxl.load_workbook(dest)
    yellow = "FFFF00"
    assert result["US"].cell(2, 1).fill.fgColor.rgb[-6:] == yellow
    assert result["AU"].cell(2, 1).fill.fgColor.rgb[-6:] == yellow
    assert result["AU"].cell(3, 1).fill.fgColor.rgb[-6:] == yellow
    # Row 3 doesn't exist on the US sheet — must not error or grow the sheet.
    assert result["US"].max_row == 2


def test_load_xls_as_workbook_copies_all_sheets(monkeypatch):
    """Regression test for the data-loss bug: _load_xls_as_workbook used to read
    only xlrd sheet index 0, silently dropping every row on later sheets of a
    multi-sheet .xls — even though parse_excel.py's actual row-matching reads
    all sheets, so those rows matched a region but then vanished from the
    highlighted output file."""
    from highlight_excel import _load_xls_as_workbook
    import xlrd

    class _FakeCell:
        def __init__(self, value):
            self.value = value
            self.ctype = xlrd.XL_CELL_TEXT

    class _FakeSheet:
        def __init__(self, name, rows):
            self.name = name
            self._rows = rows
            self.nrows = len(rows)
            self.ncols = len(rows[0]) if rows else 0

        def cell(self, r, c):
            return _FakeCell(self._rows[r][c])

    class _FakeBook:
        def __init__(self, sheets):
            self._sheets = sheets
            self.nsheets = len(sheets)

        def sheet_by_index(self, i):
            return self._sheets[i]

    sheet0 = _FakeSheet("US", [["fc", "fba"], ["BNA1", "FBA001"]])
    sheet1 = _FakeSheet("AU", [["fc", "fba"], ["SYD1", "FBA002"]])
    monkeypatch.setattr(xlrd, "open_workbook", lambda path: _FakeBook([sheet0, sheet1]))

    wb = _load_xls_as_workbook("fake.xls")

    assert [ws.title for ws in wb.worksheets] == ["US", "AU"]
    assert wb["US"].cell(2, 2).value == "FBA001"
    assert wb["AU"].cell(2, 2).value == "FBA002"
