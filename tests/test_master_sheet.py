import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import pytest

from master_sheet import (
    MASTER_SHEET_COLUMNS,
    load_master_sheet,
    save_master_sheet,
    populate_from_input,
    run_update_master_sheet,
    format_update_master_sheet_summary,
    sync_delivered_status,
)

CONTEXT_CONFIG = {
    "column_fc_code": 3,
    "column_name": 1,
    "column_ctns": 5,
    "column_shipping_way": 6,
    "column_notes": 10,
}


def _write_input_sheet(tmp_config, filename="shipments.xlsx"):
    import openpyxl as xl
    from pathlib import Path

    wb = xl.Workbook()
    ws = wb.active
    ws.append(["SYSTEM NO", "Order No", "ITEMS", "DESTINATION", "FBA ID",
               "NO OF CTNS", "SHIPPING WAY", "TRACKING NUMBERS", "CARRIER", "ETD", None])
    ws.append(["A1", "Widget Variety Pack", None, "BNA6", "FBA_CL1",
               9, "express", "1ZCL001", "UPS", None, "delivered on 2026.02.24"])
    ws.append(["A2", "Pimple Patches", None, "YVR2", "FBA_CL2",
               4, "C-SEA", "1ZCL002", "UPS", None, None])
    path = Path(tmp_config["input_folder"]) / filename
    wb.save(path)
    tmp_config.update(CONTEXT_CONFIG)
    return tmp_config


@pytest.mark.unit
def test_master_sheet_columns_order():
    assert MASTER_SHEET_COLUMNS == [
        "Tracking Status", "Delivery Date Status", "Tracking Number", "Carrier",
        "FBA ID", "Shipment Name", "Destination", "Ctns", "Shipping Way",
        "Notes (source)", "Label Created Date", "Expected Delivery Date",
        "Current Status", "Last Checked", "Region", "Workflow ID",
        "Delivery Window Start", "Delivery Window End", "Delivery Window Last Checked",
    ]


@pytest.mark.unit
def test_load_master_sheet_missing_file_returns_empty_dict(tmp_path):
    path = str(tmp_path / "does_not_exist.xlsx")
    assert load_master_sheet(path) == {}


@pytest.mark.unit
def test_save_then_load_round_trip(tmp_path):
    path = str(tmp_path / "master.xlsx")
    sheet = {
        "FBA001": {
            "tracking_status": "pending",
            "delivery_date_status": "pending",
            "tracking": "1Z001",
            "carrier": "UPS",
            "fba_id": "FBA001",
            "name": "Widget Pack",
            "destination": "ORF2",
            "ctns": 9,
            "shipping_way": "express",
            "notes": "",
            "label_created_date": "2026-06-01",
            "expected_delivery_date": "2026-06-10",
            "status": "In Transit",
            "last_checked": "2026-06-05 10:00",
            "region": "US",
            "workflow_id": "wf-abc-123",
            "delivery_window_start": "",
            "delivery_window_end": "",
            "delivery_window_last_checked": "",
        },
    }
    save_master_sheet(path, sheet)
    loaded = load_master_sheet(path)
    assert loaded == sheet


@pytest.mark.unit
def test_save_master_sheet_writes_columns_in_order(tmp_path):
    path = str(tmp_path / "master.xlsx")
    sheet = {
        "FBA001": {
            "tracking_status": "updated", "delivery_date_status": "pending",
            "tracking": "1Z001", "carrier": "UPS", "fba_id": "FBA001",
            "name": "Widget", "destination": "ORF2", "ctns": 1,
            "shipping_way": "express", "notes": "", "label_created_date": "",
            "expected_delivery_date": "", "status": "Delivered",
            "last_checked": "2026-06-05 10:00", "region": "US",
            "workflow_id": "wf-abc-123",
        },
    }
    save_master_sheet(path, sheet)

    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == MASTER_SHEET_COLUMNS
    data_row = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
    # openpyxl reads back an empty-string cell as None -- load_master_sheet
    # normalizes this back to "" for callers; this test checks raw cell values.
    assert data_row == [
        "updated", "pending", "1Z001", "UPS", "FBA001", "Widget", "ORF2", 1,
        "express", None, None, None, "Delivered", "2026-06-05 10:00", "US", "wf-abc-123",
        None, None, None,
    ]


@pytest.mark.unit
def test_save_master_sheet_overwrites_existing_row_for_same_fba_id(tmp_path):
    path = str(tmp_path / "master.xlsx")
    first = {
        "FBA001": {
            "tracking_status": "pending", "delivery_date_status": "pending",
            "tracking": "1Z001", "carrier": "UPS", "fba_id": "FBA001",
            "name": "Widget", "destination": "ORF2", "ctns": 1,
            "shipping_way": "express", "notes": "", "label_created_date": "",
            "expected_delivery_date": "", "status": "In Transit",
            "last_checked": "2026-06-01 09:00", "region": "US",
            "workflow_id": "",
        },
    }
    save_master_sheet(path, first)

    updated = load_master_sheet(path)
    updated["FBA001"]["tracking_status"] = "updated"
    updated["FBA001"]["workflow_id"] = "wf-xyz"
    save_master_sheet(path, updated)

    reloaded = load_master_sheet(path)
    assert len(reloaded) == 1
    assert reloaded["FBA001"]["tracking_status"] == "updated"
    assert reloaded["FBA001"]["workflow_id"] == "wf-xyz"


@pytest.mark.unit
def test_save_master_sheet_multiple_rows_sorted_by_fba_id(tmp_path):
    path = str(tmp_path / "master.xlsx")

    def entry(fba_id):
        return {
            "tracking_status": "pending", "delivery_date_status": "pending",
            "tracking": "1Z", "carrier": "UPS", "fba_id": fba_id,
            "name": "", "destination": "", "ctns": "", "shipping_way": "",
            "notes": "", "label_created_date": "", "expected_delivery_date": "",
            "status": "", "last_checked": "", "region": "US", "workflow_id": "",
        }

    sheet = {"FBA002": entry("FBA002"), "FBA001": entry("FBA001")}
    save_master_sheet(path, sheet)

    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    fba_ids = [row[4] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert fba_ids == ["FBA001", "FBA002"]


# --- populate_from_input -----------------------------------------------------

@pytest.mark.unit
def test_populate_from_input_creates_pending_rows_for_new_shipments(tmp_config):
    tmp_config = _write_input_sheet(tmp_config)
    sheet = populate_from_input(tmp_config, {})

    assert set(sheet.keys()) == {"FBA_CL1", "FBA_CL2"}
    row = sheet["FBA_CL1"]
    assert row["tracking_status"] == "pending"
    assert row["delivery_date_status"] == "pending"
    assert row["workflow_id"] == ""
    assert row["tracking"] == "1ZCL001"
    assert row["carrier"] == "UPS"
    assert row["fba_id"] == "FBA_CL1"
    assert row["name"] == "Widget Variety Pack"
    assert row["destination"] == "BNA6"
    assert row["ctns"] == 9
    assert row["shipping_way"] == "express"
    assert row["notes"] == "delivered on 2026.02.24"
    assert row["region"] == "US"
    # Not yet checked against a carrier -- these stay blank at population time.
    assert row["label_created_date"] == ""
    assert row["expected_delivery_date"] == ""
    assert row["status"] == ""
    assert row["last_checked"] == ""


@pytest.mark.unit
def test_populate_from_input_preserves_status_fields_for_existing_shipment(tmp_config):
    tmp_config = _write_input_sheet(tmp_config)
    existing = {
        "FBA_CL1": {
            "tracking_status": "updated", "delivery_date_status": "updated",
            "tracking": "1ZCL001", "carrier": "UPS", "fba_id": "FBA_CL1",
            "name": "Widget Variety Pack", "destination": "BNA6", "ctns": 9,
            "shipping_way": "express", "notes": "delivered on 2026.02.24",
            "label_created_date": "2026-02-10", "expected_delivery_date": "2026-02-24",
            "status": "Delivered", "last_checked": "2026-02-24 09:00",
            "region": "US", "workflow_id": "wf-already-known",
        },
    }
    sheet = populate_from_input(tmp_config, existing)

    row = sheet["FBA_CL1"]
    assert row["tracking_status"] == "updated"
    assert row["delivery_date_status"] == "updated"
    assert row["workflow_id"] == "wf-already-known"
    assert row["label_created_date"] == "2026-02-10"
    assert row["expected_delivery_date"] == "2026-02-24"
    assert row["status"] == "Delivered"
    assert row["last_checked"] == "2026-02-24 09:00"
    # New shipment from the input still gets added alongside the preserved one.
    assert "FBA_CL2" in sheet


@pytest.mark.unit
def test_populate_from_input_refreshes_source_fields_for_existing_shipment(tmp_config):
    tmp_config = _write_input_sheet(tmp_config)
    existing = {
        "FBA_CL1": {
            "tracking_status": "updated", "delivery_date_status": "pending",
            "tracking": "OLD-TRACKING", "carrier": "OLD-CARRIER", "fba_id": "FBA_CL1",
            "name": "Old Name", "destination": "OLD1", "ctns": 1,
            "shipping_way": "old-way", "notes": "old note",
            "label_created_date": "", "expected_delivery_date": "",
            "status": "", "last_checked": "", "region": "US", "workflow_id": "",
        },
    }
    sheet = populate_from_input(tmp_config, existing)

    row = sheet["FBA_CL1"]
    assert row["tracking"] == "1ZCL001"
    assert row["carrier"] == "UPS"
    assert row["name"] == "Widget Variety Pack"
    assert row["destination"] == "BNA6"
    assert row["ctns"] == 9
    assert row["shipping_way"] == "express"
    assert row["notes"] == "delivered on 2026.02.24"
    # Status field untouched by the refresh even though source fields changed.
    assert row["tracking_status"] == "updated"


# --- run_update_master_sheet / format_update_master_sheet_summary ------------

@pytest.mark.unit
def test_run_update_master_sheet_populates_saves_and_reports_counts(tmp_config):
    tmp_config = _write_input_sheet(tmp_config)
    from pathlib import Path
    tmp_config["master_sheet_path"] = str(Path(tmp_config["logs_folder"]) / "master.xlsx")
    tmp_config["tracking_status_cache"] = str(Path(tmp_config["logs_folder"]) / "does_not_exist.xlsx")

    result = run_update_master_sheet(tmp_config)

    assert result["total"] == 2
    assert result["new"] == 2
    assert result["path"] == tmp_config["master_sheet_path"]
    assert load_master_sheet(tmp_config["master_sheet_path"]) != {}


@pytest.mark.unit
def test_run_update_master_sheet_second_run_reports_zero_new(tmp_config):
    tmp_config = _write_input_sheet(tmp_config)
    from pathlib import Path
    tmp_config["master_sheet_path"] = str(Path(tmp_config["logs_folder"]) / "master.xlsx")
    tmp_config["tracking_status_cache"] = str(Path(tmp_config["logs_folder"]) / "does_not_exist.xlsx")

    run_update_master_sheet(tmp_config)
    result = run_update_master_sheet(tmp_config)

    assert result["total"] == 2
    assert result["new"] == 0


@pytest.mark.unit
def test_format_update_master_sheet_summary_includes_counts_and_path():
    text = format_update_master_sheet_summary({"total": 448, "new": 12, "path": "logs/shipment_tracking_master.xlsx"})
    assert "448" in text
    assert "12" in text
    assert "logs/shipment_tracking_master.xlsx" in text


# --- sync_delivered_status -----------------------------------------------------

def _sheet_row(tracking="1Z001", notes="", tracking_status="pending", delivery_date_status="pending"):
    return {
        "tracking": tracking, "notes": notes,
        "tracking_status": tracking_status, "delivery_date_status": delivery_date_status,
    }


@pytest.mark.unit
def test_sync_delivered_status_from_notes():
    sheet = {"FBA001": _sheet_row(notes="delivered on 2026.02.24")}
    result = sync_delivered_status(sheet, tracking_cache={})

    assert result["FBA001"]["tracking_status"] == "Delivered"
    assert result["FBA001"]["delivery_date_status"] == "Delivered"


@pytest.mark.unit
def test_sync_delivered_status_from_tracking_cache():
    sheet = {"FBA001": _sheet_row(tracking="1Z001", notes="waiting for UPS updates")}
    tracking_cache = {"1Z001": {"status": "Delivered"}}
    result = sync_delivered_status(sheet, tracking_cache)

    assert result["FBA001"]["tracking_status"] == "Delivered"
    assert result["FBA001"]["delivery_date_status"] == "Delivered"


@pytest.mark.unit
def test_sync_delivered_status_leaves_undelivered_shipments_alone():
    sheet = {"FBA001": _sheet_row(tracking="1Z001", notes="waiting for UPS updates")}
    tracking_cache = {"1Z001": {"status": "In Transit"}}
    result = sync_delivered_status(sheet, tracking_cache)

    assert result["FBA001"]["tracking_status"] == "pending"
    assert result["FBA001"]["delivery_date_status"] == "pending"


@pytest.mark.unit
def test_sync_delivered_status_overrides_updated_not_just_pending():
    sheet = {"FBA001": _sheet_row(notes="delivered on 2026.02.24", tracking_status="updated", delivery_date_status="updated")}
    result = sync_delivered_status(sheet, tracking_cache={})

    assert result["FBA001"]["tracking_status"] == "Delivered"
    assert result["FBA001"]["delivery_date_status"] == "Delivered"


@pytest.mark.unit
def test_sync_delivered_status_does_not_mutate_input():
    sheet = {"FBA001": _sheet_row(notes="delivered on 2026.02.24")}
    sync_delivered_status(sheet, tracking_cache={})

    assert sheet["FBA001"]["tracking_status"] == "pending"


@pytest.mark.unit
def test_sync_delivered_status_missing_tracking_in_cache_does_not_crash():
    sheet = {"FBA001": _sheet_row(tracking="1Z_NOT_IN_CACHE", notes="")}
    result = sync_delivered_status(sheet, tracking_cache={})

    assert result["FBA001"]["tracking_status"] == "pending"


# --- delivery window fields --------------------------------------------------

@pytest.mark.unit
def test_master_sheet_columns_include_delivery_window_fields():
    assert "Delivery Window Start" in MASTER_SHEET_COLUMNS
    assert "Delivery Window End" in MASTER_SHEET_COLUMNS
    assert "Delivery Window Last Checked" in MASTER_SHEET_COLUMNS


@pytest.mark.unit
def test_save_and_load_round_trips_delivery_window_fields(tmp_path):
    path = str(tmp_path / "master.xlsx")
    sheet = {
        "FBA001": {
            "fba_id": "FBA001", "tracking": "1Z001", "carrier": "UPS",
            "name": "", "destination": "", "ctns": "", "shipping_way": "",
            "notes": "", "region": "US", "tracking_status": "pending",
            "delivery_date_status": "pending", "label_created_date": "",
            "expected_delivery_date": "", "status": "", "last_checked": "",
            "workflow_id": "wf-1",
            "delivery_window_start": "2026-09-06",
            "delivery_window_end": "2026-09-12",
            "delivery_window_last_checked": "2026-08-30 22:00",
        }
    }
    save_master_sheet(path, sheet)
    loaded = load_master_sheet(path)
    assert loaded["FBA001"]["delivery_window_start"] == "2026-09-06"
    assert loaded["FBA001"]["delivery_window_end"] == "2026-09-12"
    assert loaded["FBA001"]["delivery_window_last_checked"] == "2026-08-30 22:00"


@pytest.mark.unit
def test_populate_from_input_new_row_has_blank_delivery_window_fields(tmp_config):
    """Extends the existing test_populate_from_input_creates_pending_rows_for_new_shipments
    (tests/test_master_sheet.py:171) pattern -- reuses its own _write_input_sheet helper,
    not a mock, matching how every other populate_from_input test in this file works."""
    tmp_config = _write_input_sheet(tmp_config)
    sheet = populate_from_input(tmp_config, {})
    row = sheet["FBA_CL1"]
    assert row["delivery_window_start"] == ""
    assert row["delivery_window_end"] == ""
    assert row["delivery_window_last_checked"] == ""
