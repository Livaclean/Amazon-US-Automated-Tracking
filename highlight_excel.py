# highlight_excel.py
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill


HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")  # yellow


def highlight_and_save(src_path: str, dest_path: str, updated_rows: set) -> str:
    """
    Opens src_path Excel file, applies yellow highlight to all cells in updated_rows
    (1-indexed row numbers), and saves to dest_path.

    Output is always .xlsx. If dest_path has a different extension, it is changed to .xlsx.
    For .xls source files, the data is converted to openpyxl format first.
    Returns the actual path written.

    Args:
        src_path: Path to original Excel file (.xlsx or .xls)
        dest_path: Desired output path
        updated_rows: Set of 1-indexed row numbers to highlight
    """
    src = Path(src_path)
    dest = Path(dest_path)

    # Always output as .xlsx (openpyxl write-only supports xlsx)
    if dest.suffix.lower() != ".xlsx":
        dest = dest.with_suffix(".xlsx")

    if src.suffix.lower() == ".xls":
        wb = _load_xls_as_workbook(src_path)
    else:
        wb = openpyxl.load_workbook(src_path)

    # row_num is a 1-indexed row within whatever sheet it came from — parse_excel.py
    # numbers rows per-sheet, not globally, so the same row_num can point at
    # unrelated data on a different sheet. Highlighting every sheet that has a
    # row there is a deliberate over-inclusive tradeoff: it guarantees the sheet
    # that was actually updated gets marked (no silent misses), at the cost of
    # occasionally highlighting an unrelated row on another sheet at the same index.
    for ws in wb.worksheets:
        max_col = ws.max_column or 1
        for row_num in updated_rows:
            if row_num > ws.max_row:
                continue
            for col in range(1, max_col + 1):
                ws.cell(row=row_num, column=col).fill = HIGHLIGHT_FILL

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dest))
    return str(dest)


def _load_xls_as_workbook(xls_path: str) -> openpyxl.Workbook:
    """Reads every sheet of a .xls file via xlrd and returns an openpyxl Workbook
    with the same sheets, in order. Mirrors parse_excel.py's all-sheets reading —
    copying only the first sheet would silently drop every row on later sheets."""
    import xlrd
    xls_wb = xlrd.open_workbook(xls_path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the blank placeholder sheet; real ones added below
    for sheet_idx in range(xls_wb.nsheets):
        xls_sheet = xls_wb.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(title=xls_sheet.name)
        for r in range(xls_sheet.nrows):
            row_data = []
            for c in range(xls_sheet.ncols):
                cell = xls_sheet.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_NUMBER:
                    val = cell.value
                    row_data.append(int(val) if val == int(val) else val)
                else:
                    row_data.append(cell.value)
            ws.append(row_data)
    return wb
